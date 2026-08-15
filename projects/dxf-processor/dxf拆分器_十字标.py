#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DXF批量拆分器 v5.37z5 - 外框剥离双模式可选版
修复：
  1) 保留 v5.37z4 保守外框检测（安全，不易拆散零件）
  2) 新增 v5.37j 精准外框检测（能剥钢模/排版图外框，可选启用）
  3) 外框剥离改为用户自选：0=不剥 1=保守 2=精准 3=精准+自定义阈值
  4) 保留所有文字提取修复（保留空格T格式、防#号误材质、序号补充等）
"""

import ezdxf
import math
import re
import shutil
import os
import logging
import traceback
import tempfile
import zipfile
import json
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from collections import defaultdict, Counter
from ezdxf import recover
from ezdxf.addons import Importer
from ezdxf.math import Matrix44
from datetime import datetime

try:
    import rarfile
    RAR_AVAILABLE = True
except ImportError:
    RAR_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠ 未安装 openpyxl，将使用 CSV 格式生成导入清单")
    print("  安装命令: pip install openpyxl")

logging.basicConfig(
    level=logging.WARNING,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

CONFIG_FILE = Path(__file__).parent / '.dxf_splitter.json'
SPLITTER_CONFIG_FILE = Path(__file__).parent / 'splitter_config.json'


def load_last_output() -> str:
    try:
        if CONFIG_FILE.exists():
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f).get('last_output_dir', '')
    except Exception:
        pass
    return ''


def save_last_output(path: str):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump({'last_output_dir': str(path)}, f, ensure_ascii=False)
    except Exception:
        pass


def load_splitter_config() -> dict:
    defaults = {
        'material_map': {},
        'process_extra': {},
        'surface_extra': {},
        'gap_tolerance': 0.5,
        'cluster_margin': 5.0,
        'min_part_size': 5.0,
    }
    try:
        if SPLITTER_CONFIG_FILE.exists():
            with open(SPLITTER_CONFIG_FILE, 'r', encoding='utf-8') as f:
                loaded = json.load(f)
                for k, v in defaults.items():
                    if k not in loaded:
                        loaded[k] = v
                return loaded
    except Exception:
        pass
    return defaults


def save_splitter_config(cfg: dict):
    try:
        with open(SPLITTER_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def edit_config_interactive(cfg: dict) -> dict:
    # 硬编码对照表（用于重复检测提示）
    hardcoded_materials = {
        '235': 'Q235', 'Q235': 'Q235', 'Q235B': 'Q235',
        '355': 'Q355', 'Q355': 'Q355', 'Q355B': 'Q355',
        '420': 'Q420', 'Q420': 'Q420', 'Q420B': 'Q420',
        '304': '304', '316': '316', '316L': '316L',
        '201': '201', 'Q201': 'Q201',
        '45': '45#', '45#': '45#', '40CR': '40Cr',
    }
    hardcoded_process = {
        'symmetry': ['对称', '正反', '正反面', '双面'],
        'bevel': ['坡口', '打坡口', '破口', '剖口'],
        'silk': ['钻孔', '打孔', '攻丝', '攻牙', '套丝'],
        'bend': ['折弯', '弯曲'],
        'cut': ['锯断', '切断', '割断'],
    }
    hardcoded_surface = {
        '镀锌': '镀锌', '热镀锌': '镀锌', '冷镀锌': '镀锌',
        '喷漆': '喷漆', '喷涂': '喷漆', '喷塑': '喷塑',
        '喷砂': '喷砂', '打砂': '喷砂', '抛丸': '喷砂', '喷丸': '喷砂',
    }
    modified = False  # [优化] 跟踪是否有修改
    while True:
        print("\n" + "="*50)
        print("【运行时配置编辑器】")
        print("="*50)
        print(f"  1. 材质索引 (当前{len(cfg.get('material_map', {}))}条自定义)")
        print(f"  2. 加工关键词 (当前{len(cfg.get('process_extra', {}))}条自定义)")
        print(f"  3. 表面状态 (当前{len(cfg.get('surface_extra', {}))}条自定义)")
        print(f"  4. 全局参数 (gap={cfg.get('gap_tolerance', 0.5)}, cluster={cfg.get('cluster_margin', 5.0)}, min_size={cfg.get('min_part_size', 5.0)}, idx_max={cfg.get('idx_search_max', 300)})")
        print("  5. 查看当前完整配置")
        print("  6. 查看所有已生效条目（含硬编码）")
        print("  0. 保存并退出")
        print("="*50)
        choice = input("选择: ").strip()
        if choice == '1':
            print("\n【材质】硬编码已有: Q235, Q355, Q420, 304, 316, 316L, 201, 45#, 40Cr")
            print("当前自定义材质:")
            for k, v in cfg.get('material_map', {}).items():
                print(f"  {k} -> {v}")
            raw = input("\n输入新增材质（格式: 缩写=标准名, 如 6061=6061），多个用逗号分隔，输入 -键名 删除）: ").strip()
            if raw:
                # [修复] 中文逗号转英文逗号，防止输入法切换导致解析错误
                raw = raw.replace('，', ',')
                for item in raw.split(','):
                    item = item.strip()
                    if not item:
                        continue
                    # 删除模式: -键名
                    if item.startswith('-'):
                        del_key = item[1:].strip()
                        if del_key in cfg.get('material_map', {}):
                            del_val = cfg['material_map'][del_key]
                            del cfg['material_map'][del_key]
                            modified = True
                            print(f"  ✅ 已删除: {del_key} -> {del_val}")
                        else:
                            print(f"  ⚠ 未找到: {del_key}")
                        continue
                    # 新增模式: 键=值
                    if '=' in item:
                        a, b = item.split('=', 1)
                        a, b = a.strip(), b.strip()
                        # 重复检测（统一转大写比较）
                        a_upper = a.upper()
                        matched_key = None
                        for hk in hardcoded_materials:
                            if hk.upper() == a_upper:
                                matched_key = hk
                                break
                        if matched_key:
                            print(f"  ⚠ 跳过: {a} 已在硬编码中 ({hardcoded_materials[matched_key]})")
                            continue
                        if a in cfg.get('material_map', {}):
                            print(f"  ⚠ 跳过: {a} 已在自定义中 ({cfg['material_map'][a]})")
                            continue
                        cfg.setdefault('material_map', {})[a] = b
                        modified = True
                        print(f"  ✅ 已添加: {a} -> {b}")
        elif choice == '2':
            print("\n【加工】硬编码已有: 对称/正反(D), 坡口(P), 钻孔/攻丝(S), 折弯(W), 锯断(J)")
            print("当前自定义加工关键词:")
            for k, v in cfg.get('process_extra', {}).items():
                print(f"  {k}: {v}")
            print("\n加工类别: symmetry(D), bevel(P), silk(S), bend(W), cut(J)")
            cat_raw = input("选择类别 [symmetry/bevel/silk/bend/cut 或 D/P/S/W/J]: ").strip().lower()
            # 支持代码位自动映射
            code_map = {'d': 'symmetry', 'p': 'bevel', 's': 'silk', 'w': 'bend', 'j': 'cut'}
            cat = code_map.get(cat_raw, cat_raw)
            raw = input("输入新增关键词（多个用逗号分隔，输入 -关键词 删除）: ").strip()
            if cat and raw:
                raw = raw.replace('，', ',')
                cfg.setdefault('process_extra', {}).setdefault(cat, [])
                added = 0
                deleted = 0
                for kw in raw.split(','):
                    kw = kw.strip()
                    if not kw:
                        continue
                    # 删除模式: -关键词
                    if kw.startswith('-'):
                        del_kw = kw[1:].strip()
                        if del_kw in cfg['process_extra'].get(cat, []):
                            cfg['process_extra'][cat].remove(del_kw)
                            deleted += 1
                            modified = True
                            print(f"  ✅ 已删除: {del_kw} 从 {cat}")
                        else:
                            print(f"  ⚠ 未找到: {del_kw} 在 {cat} 中")
                        continue
                    # 新增模式
                    if cat in hardcoded_process and kw in hardcoded_process[cat]:
                        print(f"  ⚠ 跳过: {kw} 已在硬编码 {cat} 中")
                        continue
                    if kw in cfg['process_extra'].get(cat, []):
                        print(f"  ⚠ 跳过: {kw} 已在自定义 {cat} 中")
                        continue
                    cfg['process_extra'][cat].append(kw)
                    added += 1
                    modified = True
                if added:
                    print(f"  ✅ 已添加 {added} 条到 {cat}")
                if deleted:
                    print(f"  ✅ 已删除 {deleted} 条从 {cat}")
        elif choice == '3':
            print("\n【表面状态】硬编码已有: 镀锌/热镀锌/冷镀锌, 喷漆/喷涂/喷塑, 喷砂/打砂/抛丸/喷丸")
            print("当前自定义表面状态:")
            for k, v in cfg.get('surface_extra', {}).items():
                print(f"  {k} -> {v}")
            raw = input("\n输入新增表面状态（格式: 图纸写法=文件夹后缀, 如 氧化=氧化，输入 -键名 删除）: ").strip()
            if raw:
                # 删除模式: -键名
                if raw.startswith('-'):
                    del_key = raw[1:].strip()
                    if del_key in cfg.get('surface_extra', {}):
                        del_val = cfg['surface_extra'][del_key]
                        del cfg['surface_extra'][del_key]
                        modified = True
                        print(f"  ✅ 已删除: {del_key} -> {del_val}")
                    else:
                        print(f"  ⚠ 未找到: {del_key}")
                elif '=' in raw:
                    raw = raw.replace('，', ',')
                    a, b = raw.split('=', 1)
                    a, b = a.strip(), b.strip()
                    # 重复检测
                    if a in hardcoded_surface:
                        print(f"  ⚠ 跳过: {a} 已在硬编码中 ({hardcoded_surface[a]})")
                    elif a in cfg.get('surface_extra', {}):
                        print(f"  ⚠ 跳过: {a} 已在自定义中 ({cfg['surface_extra'][a]})")
                    else:
                        cfg.setdefault('surface_extra', {})[a] = b
                        modified = True
                        print(f"  ✅ 已添加: {a} -> {b}")
        elif choice == '4':
            print(f"\n当前: gap_tolerance={cfg.get('gap_tolerance', 0.5)}, cluster_margin={cfg.get('cluster_margin', 5.0)}, min_part_size={cfg.get('min_part_size', 5.0)}")
            g = input(f"gap_tolerance (当前{cfg.get('gap_tolerance', 0.5)}): ").strip()
            c = input(f"cluster_margin (当前{cfg.get('cluster_margin', 5.0)}): ").strip()
            m = input(f"min_part_size (当前{cfg.get('min_part_size', 5.0)}): ").strip()
            if g:
                cfg['gap_tolerance'] = float(g)
                modified = True
            if c:
                cfg['cluster_margin'] = float(c)
                modified = True
            if m:
                cfg['min_part_size'] = float(m)
                modified = True
            if g or c or m:
                print("  ✅ 已更新")
            else:
                print("  无修改")
        elif choice == '5':
            print("\n当前完整配置:")
            print(json.dumps(cfg, ensure_ascii=False, indent=2))
        elif choice == '6':
            print("\n【所有已生效材质】")
            all_mats = dict(hardcoded_materials)
            all_mats.update(cfg.get('material_map', {}))
            for k, v in sorted(all_mats.items(), key=lambda x: x[0]):
                src = "硬编码" if k.upper() in hardcoded_materials else "自定义"
                print(f"  {k} -> {v} ({src})")
            print("\n【所有已生效加工关键词】")
            all_proc = {cat: list(kws) for cat, kws in hardcoded_process.items()}
            for cat, kws in cfg.get('process_extra', {}).items():
                all_proc.setdefault(cat, []).extend(kws)
            for cat, kws in sorted(all_proc.items()):
                print(f"  {cat}: {', '.join(kws)}")
            print("\n【所有已生效表面状态】")
            all_surf = dict(hardcoded_surface)
            all_surf.update(cfg.get('surface_extra', {}))
            for k, v in sorted(all_surf.items(), key=lambda x: x[0]):
                src = "硬编码" if k in hardcoded_surface else "自定义"
                print(f"  {k} -> {v} ({src})")
        elif choice == '0':
            if modified:
                save_splitter_config(cfg)
                print("\n配置已保存到 splitter_config.json")
            else:
                print("\n无修改，直接退出")
            break
    return cfg


class SmartTextExtractor:
    """智能文字提取器——宁缺勿乱，只提取明确符合逻辑的格式"""

    MATERIAL_MAP = {
        '235': 'Q235', 'Q235': 'Q235', 'Q235B': 'Q235',
        '355': 'Q355', 'Q355': 'Q355', 'Q355B': 'Q355',
        '420': 'Q420', 'Q420': 'Q420', 'Q420B': 'Q420',
        '304': '304', '316': '316', '316L': '316L',
        '201': '201', 'Q201': 'Q201',
        '45': '45#', '45#': '45#', '40CR': '40Cr',
    }

    PROCESS_KEYWORDS = {
        'symmetry': {'matches': ['对称', '正反', '正反面', '双面'], 'code': 'D'},
        'bevel': {'matches': ['坡口', '打坡口', '破口', '剖口'], 'code': 'P'},
        'silk': {'matches': ['钻孔', '打孔', '攻丝', '攻牙', '套丝'], 'code': 'S'},
        'bend': {'matches': ['折弯', '弯曲'], 'code': 'W'},
        'cut': {'matches': ['锯断', '切断', '割断'], 'code': 'J'},
    }

    SURFACE_KEYWORDS = {
        '镀锌': '镀锌', '热镀锌': '镀锌', '冷镀锌': '镀锌', ' galvanize': '镀锌',
        '喷漆': '喷漆', '喷涂': '喷漆', '喷塑': '喷塑', 'paint': '喷漆',
        '喷砂': '喷砂', '打砂': '喷砂', '抛丸': '喷砂', '喷丸': '喷砂',
    }

    def __init__(self, text_mode='compact', config=None):
        self.text_mode = text_mode  # 'compact' 或 'loose'
        # [修复] 复制类变量到实例变量，防止 update 污染类
        self.MATERIAL_MAP = dict(self.__class__.MATERIAL_MAP)
        self.PROCESS_KEYWORDS = {
            k: {'matches': list(v['matches']), 'code': v['code']}
            for k, v in self.__class__.PROCESS_KEYWORDS.items()
        }
        self.SURFACE_KEYWORDS = dict(self.__class__.SURFACE_KEYWORDS)
        # 加载外部配置覆盖默认值
        if config:
            self.MATERIAL_MAP.update(config.get('material_map', {}))
            for cat, kws in config.get('process_extra', {}).items():
                if cat in self.PROCESS_KEYWORDS and kws:
                    self.PROCESS_KEYWORDS[cat]['matches'].extend(kws)
            self.SURFACE_KEYWORDS.update(config.get('surface_extra', {}))

    def extract_all_texts(self, msp) -> Tuple[List[Dict], List[Dict], List[Dict], List[Dict], List[Dict], List[Dict]]:
        if self.text_mode == 'loose':
            return self._extract_loose(msp)
        return self._extract_compact(msp)

    def _clean_mtext(self, text: str) -> str:
        """增强版 MTEXT 清理 - 融合 v5.24.8 的完善格式码处理"""
        if not text:
            return ''
        text = re.sub(r'\\P', '\n', text)
        text = re.sub(r'\\H\d+(?:\.\d+)?[xX];', '', text)
        text = re.sub(r'\\S[^;]*;', '', text)
        text = re.sub(r'\\A\d+;', '', text)
        text = re.sub(r'\\[Cc]\d+;', '', text)
        text = re.sub(r'\\F[^;]*;', '', text)
        text = re.sub(r'\\[A-Za-z]+\d*[^;]*;', '', text)
        # [v5.37z8-修复] 清理无分号格式码如 \fSimSun|b0|i0|c134，排除{}
        # [v5.37z17-修复] 只匹配ASCII字符，防止吃掉后面的中文加工标注
        text = re.sub(r'\\[A-Za-z]+[A-Za-z0-9|._-]*', '', text)
        text = re.sub(r'[{}]', '', text)
        text = re.sub(r'\\.', '', text)
        text = re.sub(r'\|[A-Za-z0-9]+', '', text)
        return text.strip()

    def _extract_compact(self, msp) -> Tuple[List[Dict], List[Dict], List[Dict], List[Dict], List[Dict], List[Dict]]:
        attr_texts = []
        index_texts = []
        process_texts = []
        material_texts = []
        split_texts = []

        for text_type in ['TEXT', 'MTEXT']:
            for entity in msp.query(text_type):
                try:
                    if text_type == 'TEXT':
                        content_lines = [entity.dxf.text.strip()]
                        pos = (entity.dxf.insert[0], entity.dxf.insert[1])
                    else:
                        raw_text = self._clean_mtext(entity.text)
                        # [v5.37p 修复1] MTEXT 按 \\P 拆分为多行，每行独立提取
                        content_lines = [line.strip() for line in re.split(r'\\P|\r?\n', raw_text) if line.strip()]
                        pos = (entity.dxf.insert[0], entity.dxf.insert[1])

                    for content in content_lines:
                        if not content:
                            continue

                        text_upper = content.upper()
                        text_upper_nospace = text_upper.replace(' ', '')

                        # ========== 第零轮:材质标注（用去空格版本）==========
                        mat = self.parse_material(content)
                        if mat:
                            material_texts.append({
                                'content': content,
                                'pos': pos,
                                'material': mat,
                            })

                        # [v5.37z8-修复] 先删除已识别材质关键词，防止被数量正则贪婪吃掉
                        mat = self.parse_material(content)
                        mat_cleaned = text_upper_nospace
                        if mat:
                            for alias, standard in sorted(self.MATERIAL_MAP.items(), key=lambda x: -len(x[0])):
                                alias_upper = alias.upper()
                                if alias_upper in mat_cleaned:
                                    mat_cleaned = mat_cleaned.replace(alias_upper, '')
                                    break

                        # ========== 第一轮:T*格式（用去空格版本）==========
                        cleaned_upper = re.sub(r'(T\d+(?:\.\d+)?)[\*Xx×＊](\d+)', r'\1*\2', mat_cleaned)

                        # [v5.37z6 修复] 支持拼接标记如7#-2T4*88，一行多T*如15#T10*88 T8*44
                        matches = list(re.finditer(r'(?:(\d{1,2})[#＃](?:[_-]\d+)?(?:[A-Za-z0-9]+)?)?T(\d+(?:\.\d+)?)[\*Xx](\d+)', cleaned_upper))
                        if matches:
                            # 提取行内共享序号
                            line_index = None
                            for m in matches:
                                if m.group(1):
                                    line_index = m.group(1)
                                    break

                            # 行内材质/加工只添加一次
                            mat = self.parse_material(content)
                            if mat:
                                material_texts.append({'content': content, 'pos': pos, 'material': mat})
                            proc = self.parse_process_code(content)
                            if proc != 'xxxxx':
                                process_texts.append({'content': content, 'pos': pos, 'process_code': proc})

                            for m in matches:
                                idx_val = m.group(1) or line_index
                                # T*前面没有序号时，在同一行剩余部分顺手找序号
                                if not idx_val:
                                    tail = cleaned_upper[m.end():]
                                    tail_match = re.search(r'(\d{1,2})[#＃]', tail)
                                    if tail_match:
                                        idx_val = tail_match.group(1)
                                        already_exists = any(it['index'] == idx_val and
                                                           math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                                           for it in index_texts)
                                        if not already_exists:
                                            index_texts.append({
                                                'content': content, 'pos': pos, 'priority': 50,
                                                'index': idx_val,
                                            })

                                # [v5.37z6-修复] 顺手提取拼接标记如 6#-3
                                if idx_val and idx_val != 'xx':
                                    split_match = re.search(r'(?<![A-Z0-9])' + re.escape(idx_val) + r'[#＃]-(\\d+)(?![0-9#＃])', cleaned_upper)
                                    if not split_match:
                                        split_match = re.search(r'(?<![A-Z0-9])' + re.escape(idx_val) + r'[#＃]_(\\d+)(?![0-9#＃])', cleaned_upper)
                                    if split_match:
                                        sub_idx_val = f"{int(split_match.group(1)):02d}"
                                        already_split = any(st.get('index') == idx_val and st.get('sub_index') == sub_idx_val
                                                        and math.hypot(st['pos'][0]-pos[0], st['pos'][1]-pos[1]) < 1.0
                                                        for st in split_texts)
                                        if not already_split:
                                            split_texts.append({
                                                'content': content, 'pos': pos, 'priority': 55,
                                                'index': idx_val, 'sub_index': sub_idx_val,
                                            })

                                attr_texts.append({
                                    'content': content, 'pos': pos, 'priority': 100,
                                    'index': idx_val or 'xx',
                                    'thickness': float(m.group(2)),
                                    'qty': int(m.group(3)),
                                    'has_qty': True,
                                })
                            continue

                        # ========== 第二轮:T格式（保留空格，防止粘连误匹配）==========
                        # [v5.37z2 保留] 改用保留空格版本，防止"T16 12#"去空格后粘连成"T1612#"导致厚度丢失
                        # [v5.37z6 修复] 支持拼接标记如7#-2T4，尾部抓数量如T8*4
                        matches = list(re.finditer(r'(?:(\d{1,2})\s*[#＃]\s*(?:[_-]\d+)?)?\s*T\s*(\d+(?:\.\d+)?)(?![\d\.])', text_upper))
                        if matches:
                            valid_matches = [m for m in matches if float(m.group(2)) <= 100]
                            if not valid_matches:
                                matches = []
                            else:
                                matches = valid_matches
                        if matches:
                            # 行内材质/加工只添加一次
                            mat = self.parse_material(content)
                            if mat:
                                material_texts.append({'content': content, 'pos': pos, 'material': mat})
                            proc = self.parse_process_code(content)
                            if proc != 'xxxxx':
                                process_texts.append({'content': content, 'pos': pos, 'process_code': proc})

                            for match in matches:
                                idx_val = match.group(1)
                                # T前面没有序号时，在同一行剩余部分顺手找序号
                                if not idx_val:
                                    tail = text_upper[match.end():].replace(' ', '')
                                    tail_match = re.search(r'(\d{1,2})[#＃]', tail)
                                    if tail_match:
                                        idx_val = tail_match.group(1)
                                        already_exists = any(it['index'] == idx_val and
                                                           math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                                           for it in index_texts)
                                        if not already_exists:
                                            index_texts.append({
                                                'content': content, 'pos': pos, 'priority': 50,
                                                'index': idx_val,
                                            })

                                # [v5.37z6 修复] 检查尾部*数量
                                tail_qty = None
                                tail_match = re.search(r'[\*Xx×＊](\d+)', text_upper[match.end():])
                                if tail_match:
                                    tail_qty = int(tail_match.group(1))

                                attr_texts.append({
                                    'content': content, 'pos': pos, 'priority': 80,
                                    'index': idx_val or 'xx',
                                    'thickness': float(match.group(2)),
                                    'qty': tail_qty if tail_qty else 1,
                                    'has_qty': bool(tail_qty),
                                })
                            continue

                        # ========== 第三轮:拼接格式（用保留空格版本！）==========
                        idx_val = None
                        sub_idx_val = None

                        match_split = re.search(r'(?<![A-Z0-9])(\d{1,2})[#＃]_(\d+)(?![A-Z0-9])', text_upper)
                        if match_split and 'T' not in text_upper:
                            idx_val = match_split.group(1)
                            sub_idx_val = f"{int(match_split.group(2)):02d}"
                            split_texts.append({
                                'content': content, 'pos': pos, 'priority': 55,
                                'index': idx_val, 'sub_index': sub_idx_val,
                            })
                            already_exists = any(it['index'] == idx_val and
                                               math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                               for it in index_texts)
                            if not already_exists:
                                index_texts.append({
                                    'content': content, 'pos': pos, 'priority': 50,
                                    'index': idx_val,
                                })
                            continue

                        match_split2 = re.search(r'(?<![A-Z0-9])(\d{1,2})[#＃]-(\d+)(?![0-9#＃])', text_upper)
                        if match_split2:
                            # [v5.37z6 修复] 去掉 'T' not in text_upper 的整行屏蔽，7#-2T4*88 需要提取-2
                            idx_val = match_split2.group(1)
                            # [v5.37z8-保险] 序号超过99直接丢弃
                            if int(idx_val) > 99:
                                continue
                            sub_idx_val = f"{int(match_split2.group(2)):02d}"
                            split_texts.append({
                                'content': content, 'pos': pos, 'priority': 55,
                                'index': idx_val, 'sub_index': sub_idx_val,
                            })
                            already_exists = any(it['index'] == idx_val and
                                               math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                               for it in index_texts)
                            if not already_exists:
                                index_texts.append({
                                    'content': content, 'pos': pos, 'priority': 50,
                                    'index': idx_val,
                                })
                            continue

                        # [v5.37z8-修复] 新增 7号-1 格式支持
                        match_split3 = re.search(r'(?<![A-Z0-9])(\d{1,2})[号號]-(\d+)(?![0-9号號])', text_upper)
                        if match_split3:
                            idx_val = match_split3.group(1)
                            # [v5.37z8-保险] 序号超过99直接丢弃
                            if int(idx_val) > 99:
                                continue
                            sub_idx_val = f"{int(match_split3.group(2)):02d}"
                            split_texts.append({
                                'content': content, 'pos': pos, 'priority': 55,
                                'index': idx_val, 'sub_index': sub_idx_val,
                            })
                            already_exists = any(it['index'] == idx_val and
                                               math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                               for it in index_texts)
                            if not already_exists:
                                index_texts.append({
                                    'content': content, 'pos': pos, 'priority': 50,
                                    'index': idx_val,
                                })
                            continue

                        # [v5.37z8-修复] 新增 7号-1 格式支持
                        match_split3 = re.search(r'(?<![A-Z0-9])(\d{1,2})[号號]-(\d+)(?![0-9号號])', text_upper)
                        if match_split3:
                            idx_val = match_split3.group(1)
                            # [v5.37z8-保险] 序号超过99直接丢弃
                            if int(idx_val) > 99:
                                continue
                            sub_idx_val = f"{int(match_split3.group(2)):02d}"
                            split_texts.append({
                                'content': content, 'pos': pos, 'priority': 55,
                                'index': idx_val, 'sub_index': sub_idx_val,
                            })
                            already_exists = any(it['index'] == idx_val and
                                               math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                               for it in index_texts)
                            if not already_exists:
                                index_texts.append({
                                    'content': content, 'pos': pos, 'priority': 50,
                                    'index': idx_val,
                                })
                            continue

                        # ========== 第四轮:标准序号格式（用去空格版本！）==========
                        idx_val = None

                        # [v5.37z8-修复] 只匹配数字#格式（如7#, 14#），限制1-2位
                        match = re.search(r'(\d{1,2})[#＃]', text_upper_nospace)
                        if match and 'T' not in text_upper_nospace and match.group(1) not in ['', '0']:
                            idx_val = match.group(1)

                        if not idx_val:
                            match2 = re.search(r'(\d{1,2})-\d+[号號]', text_upper_nospace)
                            if match2 and 'T' not in text_upper_nospace:
                                idx_val = match2.group(1)

                        if not idx_val:
                            match3 = re.search(r'(\d{1,2})[号號]', text_upper_nospace)
                            if match3 and 'T' not in text_upper_nospace:
                                idx_val = match3.group(1)

                        if idx_val:
                            # [v5.37z8-保险] 序号超过99（3位以上）直接丢弃，防止材质/零件编号误抓
                            if int(idx_val) > 99:
                                continue
                            already_exists = any(it['index'] == idx_val and
                                               math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                               for it in index_texts)
                            if not already_exists:
                                index_texts.append({
                                    'content': content, 'pos': pos, 'priority': 50,
                                    'index': idx_val,
                                })
                            continue

                        # ========== 第五轮:独立拼接标记（用保留空格版本！）==========
                        match_ind_split = re.search(r'(?<!\d)[_-](\d{2})(?!\d)', text_upper)
                        if match_ind_split and 'T' not in text_upper:
                            split_texts.append({
                                'content': content, 'pos': pos, 'priority': 45,
                                'index': None,
                                'sub_index': match_ind_split.group(1),
                            })
                            continue

                        # ========== 第六轮:加工标注 ==========
                        process_code = self.parse_process_code(content)
                        if process_code != 'xxxxx':
                            process_texts.append({
                                'content': content,
                                'pos': pos,
                                'process_code': process_code,
                            })

                except Exception as e:
                    logger.debug(f"文字提取异常: {e}")
                    continue

        attr_texts.sort(key=lambda t: t['priority'], reverse=True)

        # [v5.37z6-修复3] 收集独立数量标注（如 *4、×5、X3 等单独文字实体）
        qty_texts = []
        for text_type in ['TEXT', 'MTEXT']:
            for entity in msp.query(text_type):
                try:
                    if text_type == 'TEXT':
                        content_lines = [entity.dxf.text.strip()]
                        pos = (entity.dxf.insert[0], entity.dxf.insert[1])
                    else:
                        raw_text = self._clean_mtext(entity.text)
                        content_lines = [line.strip() for line in re.split(r'\\P|\r?\n', raw_text) if line.strip()]
                        pos = (entity.dxf.insert[0], entity.dxf.insert[1])
                    for content in content_lines:
                        if not content:
                            continue
                        text_upper = content.upper().replace(' ', '')
                        # 独立数量：*数字、×数字、X数字（前面没有T）
                        qm = re.search(r'^[\\*Xx×＊](\d+)$', text_upper)
                        if qm and not re.search(r'T\d', text_upper):
                            qty_texts.append({
                                'content': content,
                                'pos': pos,
                                'qty': int(qm.group(1)),
                            })
                except Exception:
                    continue

        return attr_texts, index_texts, process_texts, material_texts, split_texts, qty_texts

    def parse_material(self, text: str) -> Optional[str]:
        if self.text_mode == 'loose':
            return self._parse_material_loose(text)
        return self._parse_material_compact(text)

    def _parse_material_compact(self, text: str) -> Optional[str]:
        if not text:
            return None

        # [v5.37z6 修复] 先抠掉序号部分（如7#-2, 15#, 3#_02），避免误杀
        text_clean = re.sub(r'\d+\s*[#＃]\s*(?:[_-]\d+)?', '', text)
        # 如果去掉序号后还有#号（说明是其他用途），再检查
        if '#' in text_clean or '＃' in text_clean:
            return None

        # 全角数字/字母转半角，提高匹配兼容性
        fullwidth = '０１２３４５６７８９ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ'
        halfwidth = '0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz'
        text_clean = text_clean.translate(str.maketrans(fullwidth, halfwidth))

        text_upper = text_clean.upper().replace(' ', '')
        text_lower = text_clean.lower()

        exclude_patterns = [
            r'缩尺到\d+',
            r'缩尺\d+',
            r'到\d+号',
            r'\d+号件',
        ]
        for pattern in exclude_patterns:
            if re.search(pattern, text_lower):
                return None

        # [v5.37z6-修复] 按长度降序遍历，优先匹配更长的如 Q201 而不是 201
        for alias, standard in sorted(self.MATERIAL_MAP.items(), key=lambda x: -len(x[0])):
            if alias.upper() in text_upper:
                return standard
        return None

    def _extract_loose(self, msp) -> Tuple[List[Dict], List[Dict], List[Dict], List[Dict], List[Dict], List[Dict]]:
        attr_texts = []
        index_texts = []
        process_texts = []
        material_texts = []
        split_texts = []

        for text_type in ['TEXT', 'MTEXT']:
            for entity in msp.query(text_type):
                try:
                    if text_type == 'TEXT':
                        content_lines = [entity.dxf.text.strip()]
                        pos = (entity.dxf.insert[0], entity.dxf.insert[1])
                    else:
                        raw_text = self._clean_mtext(entity.text)
                        # [松散模式-复制精确正则] MTEXT 按 \\P 拆分为多行，每行独立提取
                        content_lines = [line.strip() for line in re.split(r'\\P|\r?\n', raw_text) if line.strip()]
                        pos = (entity.dxf.insert[0], entity.dxf.insert[1])

                    for content in content_lines:
                        if not content:
                            continue

                        text_upper = content.upper()
                        text_upper_nospace = text_upper.replace(' ', '')

                        # ========== 第零轮:材质标注（用去空格版本）==========
                        mat = self.parse_material(content)
                        if mat:
                            material_texts.append({
                                'content': content,
                                'pos': pos,
                                'material': mat,
                            })

                        # [v5.37z8-修复] 先删除已识别材质关键词，防止被数量正则贪婪吃掉
                        mat = self.parse_material(content)
                        mat_cleaned = text_upper_nospace
                        if mat:
                            for alias, standard in sorted(self.MATERIAL_MAP.items(), key=lambda x: -len(x[0])):
                                alias_upper = alias.upper()
                                if alias_upper in mat_cleaned:
                                    mat_cleaned = mat_cleaned.replace(alias_upper, '')
                                    break

                        # ========== 第一轮:T*格式（用去空格版本）==========
                        cleaned_upper = re.sub(r'(T\d+(?:\.\d+)?)[\*Xx×＊](\d+)', r'\1*\2', mat_cleaned)

                        # [v5.37z6 修复] 支持拼接标记如7#-2T4*88，一行多T*如15#T10*88 T8*44
                        matches = list(re.finditer(r'(?:(\d{1,2})[#＃](?:[_-]\d+)?(?:[A-Za-z0-9]+)?)?T(\d+(?:\.\d+)?)[\*Xx](\d+)', cleaned_upper))
                        if matches:
                            # 提取行内共享序号
                            line_index = None
                            for m in matches:
                                if m.group(1):
                                    line_index = m.group(1)
                                    break

                            # 行内材质/加工只添加一次
                            mat = self.parse_material(content)
                            if mat:
                                material_texts.append({'content': content, 'pos': pos, 'material': mat})
                            proc = self.parse_process_code(content)
                            if proc != 'xxxxx':
                                process_texts.append({'content': content, 'pos': pos, 'process_code': proc})

                            for m in matches:
                                idx_val = m.group(1) or line_index
                                # T*前面没有序号时，在同一行剩余部分顺手找序号
                                if not idx_val:
                                    tail = cleaned_upper[m.end():]
                                    tail_match = re.search(r'(\d{1,2})[#＃]', tail)
                                    if tail_match:
                                        idx_val = tail_match.group(1)
                                        already_exists = any(it['index'] == idx_val and
                                                           math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                                           for it in index_texts)
                                        if not already_exists:
                                            index_texts.append({
                                                'content': content, 'pos': pos, 'priority': 50,
                                                'index': idx_val,
                                            })

                                # [松散模式-复制精确正则] 顺手提取拼接标记如 6#-3
                                if idx_val and idx_val != 'xx':
                                    split_match = re.search(r'(?<![A-Z0-9])' + re.escape(idx_val) + r'[#＃]-(\\d+)(?![0-9#＃])', cleaned_upper)
                                    if not split_match:
                                        split_match = re.search(r'(?<![A-Z0-9])' + re.escape(idx_val) + r'[#＃]_(\\d+)(?![0-9#＃])', cleaned_upper)
                                    if split_match:
                                        sub_idx_val = f"{int(split_match.group(1)):02d}"
                                        already_split = any(st.get('index') == idx_val and st.get('sub_index') == sub_idx_val
                                                        and math.hypot(st['pos'][0]-pos[0], st['pos'][1]-pos[1]) < 1.0
                                                        for st in split_texts)
                                        if not already_split:
                                            split_texts.append({
                                                'content': content, 'pos': pos, 'priority': 55,
                                                'index': idx_val, 'sub_index': sub_idx_val,
                                            })

                                attr_texts.append({
                                    'content': content, 'pos': pos, 'priority': 100,
                                    'index': idx_val or 'xx',
                                    'thickness': float(m.group(2)),
                                    'qty': int(m.group(3)),
                                    'has_qty': True,
                                })
                            continue

                        # ========== 第二轮:T格式（保留空格，防止粘连误匹配）==========
                        # [松散模式-复制精确正则] 改用保留空格版本，防止"T16 12#"去空格后粘连成"T1612#"导致厚度丢失
                        # [v5.37z6 修复] 支持拼接标记如7#-2T4，尾部抓数量如T8*4
                        matches = list(re.finditer(r'(?:(\d{1,2})\s*[#＃]\s*(?:[_-]\d+)?)?\s*T\s*(\d+(?:\.\d+)?)(?![\d\.])', text_upper))
                        if matches:
                            valid_matches = [m for m in matches if float(m.group(2)) <= 100]
                            if not valid_matches:
                                matches = []
                            else:
                                matches = valid_matches
                        if matches:
                            # 行内材质/加工只添加一次
                            mat = self.parse_material(content)
                            if mat:
                                material_texts.append({'content': content, 'pos': pos, 'material': mat})
                            proc = self.parse_process_code(content)
                            if proc != 'xxxxx':
                                process_texts.append({'content': content, 'pos': pos, 'process_code': proc})

                            for match in matches:
                                idx_val = match.group(1)
                                # T前面没有序号时，在同一行剩余部分顺手找序号
                                if not idx_val:
                                    tail = text_upper[match.end():].replace(' ', '')
                                    tail_match = re.search(r'(\d{1,2})[#＃]', tail)
                                    if tail_match:
                                        idx_val = tail_match.group(1)
                                        already_exists = any(it['index'] == idx_val and
                                                           math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                                           for it in index_texts)
                                        if not already_exists:
                                            index_texts.append({
                                                'content': content, 'pos': pos, 'priority': 50,
                                                'index': idx_val,
                                            })

                                # [v5.37z6 修复] 检查尾部*数量
                                tail_qty = None
                                tail_match = re.search(r'[\*Xx×＊](\d+)', text_upper[match.end():])
                                if tail_match:
                                    tail_qty = int(tail_match.group(1))

                                attr_texts.append({
                                    'content': content, 'pos': pos, 'priority': 80,
                                    'index': idx_val or 'xx',
                                    'thickness': float(match.group(2)),
                                    'qty': tail_qty if tail_qty else 1,
                                    'has_qty': bool(tail_qty),
                                })
                            continue

                        # ========== 第三轮:拼接格式（用保留空格版本！）==========
                        idx_val = None
                        sub_idx_val = None

                        match_split = re.search(r'(?<![A-Z0-9])(\d{1,2})[#＃]_(\d+)(?![A-Z0-9])', text_upper)
                        if match_split and 'T' not in text_upper:
                            idx_val = match_split.group(1)
                            sub_idx_val = f"{int(match_split.group(2)):02d}"
                            split_texts.append({
                                'content': content, 'pos': pos, 'priority': 55,
                                'index': idx_val, 'sub_index': sub_idx_val,
                            })
                            already_exists = any(it['index'] == idx_val and
                                               math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                               for it in index_texts)
                            if not already_exists:
                                index_texts.append({
                                    'content': content, 'pos': pos, 'priority': 50,
                                    'index': idx_val,
                                })
                            continue

                        match_split2 = re.search(r'(?<![A-Z0-9])(\d{1,2})[#＃]-(\d+)(?![0-9#＃])', text_upper)
                        if match_split2:
                            # [v5.37z6 修复] 去掉 'T' not in text_upper 的整行屏蔽，7#-2T4*88 需要提取-2
                            idx_val = match_split2.group(1)
                            # [v5.37z8-保险] 序号超过99直接丢弃
                            if int(idx_val) > 99:
                                continue
                            sub_idx_val = f"{int(match_split2.group(2)):02d}"
                            split_texts.append({
                                'content': content, 'pos': pos, 'priority': 55,
                                'index': idx_val, 'sub_index': sub_idx_val,
                            })
                            already_exists = any(it['index'] == idx_val and
                                               math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                               for it in index_texts)
                            if not already_exists:
                                index_texts.append({
                                    'content': content, 'pos': pos, 'priority': 50,
                                    'index': idx_val,
                                })
                            continue

                        # [v5.37z8-修复] 新增 7号-1 格式支持
                        match_split3 = re.search(r'(?<![A-Z0-9])(\d{1,2})[号號]-(\d+)(?![0-9号號])', text_upper)
                        if match_split3:
                            idx_val = match_split3.group(1)
                            # [v5.37z8-保险] 序号超过99直接丢弃
                            if int(idx_val) > 99:
                                continue
                            sub_idx_val = f"{int(match_split3.group(2)):02d}"
                            split_texts.append({
                                'content': content, 'pos': pos, 'priority': 55,
                                'index': idx_val, 'sub_index': sub_idx_val,
                            })
                            already_exists = any(it['index'] == idx_val and
                                               math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                               for it in index_texts)
                            if not already_exists:
                                index_texts.append({
                                    'content': content, 'pos': pos, 'priority': 50,
                                    'index': idx_val,
                                })
                            continue

                        # [v5.37z8-修复] 新增 7号-1 格式支持
                        match_split3 = re.search(r'(?<![A-Z0-9])(\d{1,2})[号號]-(\d+)(?![0-9号號])', text_upper)
                        if match_split3:
                            idx_val = match_split3.group(1)
                            # [v5.37z8-保险] 序号超过99直接丢弃
                            if int(idx_val) > 99:
                                continue
                            sub_idx_val = f"{int(match_split3.group(2)):02d}"
                            split_texts.append({
                                'content': content, 'pos': pos, 'priority': 55,
                                'index': idx_val, 'sub_index': sub_idx_val,
                            })
                            already_exists = any(it['index'] == idx_val and
                                               math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                               for it in index_texts)
                            if not already_exists:
                                index_texts.append({
                                    'content': content, 'pos': pos, 'priority': 50,
                                    'index': idx_val,
                                })
                            continue

                        # ========== 第四轮:标准序号格式（用去空格版本！）==========
                        idx_val = None

                        # [v5.37z8-修复] 只匹配数字#格式（如7#, 14#），限制1-2位
                        match = re.search(r'(\d{1,2})[#＃]', text_upper_nospace)
                        if match and 'T' not in text_upper_nospace and match.group(1) not in ['', '0']:
                            idx_val = match.group(1)

                        if not idx_val:
                            match2 = re.search(r'(\d{1,2})-\d+[号號]', text_upper_nospace)
                            if match2 and 'T' not in text_upper_nospace:
                                idx_val = match2.group(1)

                        if not idx_val:
                            match3 = re.search(r'(\d{1,2})[号號]', text_upper_nospace)
                            if match3 and 'T' not in text_upper_nospace:
                                idx_val = match3.group(1)

                        if idx_val:
                            # [v5.37z8-保险] 序号超过99（3位以上）直接丢弃，防止材质/零件编号误抓
                            if int(idx_val) > 99:
                                continue
                            already_exists = any(it['index'] == idx_val and
                                               math.hypot(it['pos'][0]-pos[0], it['pos'][1]-pos[1]) < 1.0
                                               for it in index_texts)
                            if not already_exists:
                                index_texts.append({
                                    'content': content, 'pos': pos, 'priority': 50,
                                    'index': idx_val,
                                })
                            continue

                        # ========== 第五轮:独立拼接标记（用保留空格版本！）==========
                        match_ind_split = re.search(r'(?<!\d)[_-](\d{2})(?!\d)', text_upper)
                        if match_ind_split and 'T' not in text_upper:
                            split_texts.append({
                                'content': content, 'pos': pos, 'priority': 45,
                                'index': None,
                                'sub_index': match_ind_split.group(1),
                            })
                            continue

                        # ========== 第六轮:加工标注 ==========
                        process_code = self.parse_process_code(content)
                        if process_code != 'xxxxx':
                            process_texts.append({
                                'content': content,
                                'pos': pos,
                                'process_code': process_code,
                            })

                except Exception as e:
                    logger.debug(f"文字提取异常: {e}")
                    continue

        attr_texts.sort(key=lambda t: t['priority'], reverse=True)

        # [松散模式-复制精确正则] 收集独立数量标注（如 *4、×5、X3 等单独文字实体）
        qty_texts = []
        for text_type in ['TEXT', 'MTEXT']:
            for entity in msp.query(text_type):
                try:
                    if text_type == 'TEXT':
                        content_lines = [entity.dxf.text.strip()]
                        pos = (entity.dxf.insert[0], entity.dxf.insert[1])
                    else:
                        raw_text = self._clean_mtext(entity.text)
                        content_lines = [line.strip() for line in re.split(r'\\P|\r?\n', raw_text) if line.strip()]
                        pos = (entity.dxf.insert[0], entity.dxf.insert[1])
                    for content in content_lines:
                        if not content:
                            continue
                        text_upper = content.upper().replace(' ', '')
                        # 独立数量：*数字、×数字、X数字（前面没有T）
                        qm = re.search(r'^[\\*Xx×＊](\d+)$', text_upper)
                        if qm and not re.search(r'T\d', text_upper):
                            qty_texts.append({
                                'content': content,
                                'pos': pos,
                                'qty': int(qm.group(1)),
                            })
                except Exception:
                    continue

        return attr_texts, index_texts, process_texts, material_texts, split_texts, qty_texts


    def _parse_material_loose(self, text: str) -> Optional[str]:
        if not text:
            return None

        # [v5.37z6 修复] 先抠掉序号部分（如7#-2, 15#, 3#_02），避免误杀
        text_clean = re.sub(r'\d+\s*[#＃]\s*(?:[_-]\d+)?', '', text)
        # 如果去掉序号后还有#号（说明是其他用途），再检查
        if '#' in text_clean or '＃' in text_clean:
            return None

        # 全角数字/字母转半角，提高匹配兼容性
        fullwidth = '０１２３４５６７８９ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ'
        halfwidth = '0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz'
        text_clean = text_clean.translate(str.maketrans(fullwidth, halfwidth))

        text_upper = text_clean.upper().replace(' ', '')
        text_lower = text_clean.lower()

        exclude_patterns = [
            r'缩尺到\d+',
            r'缩尺\d+',
            r'到\d+号',
            r'\d+号件',
        ]
        for pattern in exclude_patterns:
            if re.search(pattern, text_lower):
                return None

        # [松散模式-复制精确正则] 按长度降序遍历，优先匹配更长的如 Q201 而不是 201
        for alias, standard in sorted(self.MATERIAL_MAP.items(), key=lambda x: -len(x[0])):
            if alias.upper() in text_upper:
                return standard
        return None


    def parse_process_code(self, text: str) -> str:
        text_lower = text.lower()
        code = ['x'] * 5
        for idx, (key, config) in enumerate(self.PROCESS_KEYWORDS.items()):
            for keyword in config['matches']:
                if keyword in text_lower:
                    code[idx] = config['code']
                    break
        return ''.join(code)

    def parse_surface(self, text: str) -> str:
        if not text:
            return ''
        text_lower = text.lower()
        for keyword, surface in self.SURFACE_KEYWORDS.items():
            if keyword.lower() in text_lower:
                return surface
        return ''


class UniversalPartSplitter:
    def __init__(self, min_part_size=5.0, gap_tolerance=0.5, cluster_margin=5.0, explode_blocks=True,
                 frame_mode='conservative', frame_threshold=300, config=None):
        self.min_part_size = min_part_size
        self.gap_tolerance = gap_tolerance
        self.cluster_margin = cluster_margin
        self.explode_blocks = explode_blocks
        self.frame_mode = frame_mode          # 'none', 'conservative', 'precise'
        self.frame_threshold = frame_threshold  # 精准模式最小外框对角线
        self.drawing_no = ""
        self.piece_count = 1
        self.full_prefix = ""
        self.stats = {'files': 0, 'parts': 0, 'matched': 0, 'unmatched': 0, 'frames': 0}
        self.frame_entities = []
        self.text_extractor = SmartTextExtractor(config=config)
        self.config = config or {}

    def select_parameters_interactive(self):
        # 加载外部配置
        cfg = load_splitter_config()
        # 询问是否进入配置编辑（默认跳过，按9才进入）
        print("\n" + "="*60)
        print("【配置管理】 直接回车跳过 | 按 9 进入编辑器")
        print("="*60)
        edit_choice = input("[回车跳过/9编辑]: ").strip()
        if edit_choice == '9':
            cfg = edit_config_interactive(cfg)
        # 用配置覆盖实例参数
        self.gap_tolerance = cfg.get('gap_tolerance', self.gap_tolerance)
        self.cluster_margin = cfg.get('cluster_margin', self.cluster_margin)
        self.min_part_size = cfg.get('min_part_size', self.min_part_size)
        self.config = cfg

        # [v5.37z8-新增] 临时参数：小零件序号搜索距离覆盖（不保存到配置）
        print("\n【临时参数】直接回车跳过 | 输入数字覆盖本次运行的序号搜索阈值")
        temp_idx = input("小零件序号最大搜索距离mm[默认 part_diag*0.8]: ").strip()
        if temp_idx:
            self.config['idx_search_max'] = float(temp_idx)
            print(f"   本次运行覆盖: 小零件序号阈值 = {temp_idx}mm")
        else:
            print("   使用默认阈值: part_diag * 0.8")

        print("\n" + "="*60)
        print("【文字提取模式选择】")
        print('='*60)
        print("  1. 【紧凑精确】规范制图，零件间隔大，文字紧凑")
        print("     - 正则精确（支持拼接标记如7#-2T4*88）")
        print("     - 搜索半径小，宁缺勿乱，防止远处文字误分配")
        print("     - 独占策略：厚度/序号/材质/加工 抢完为止")
        print("  2. 【宽松覆盖】混乱图纸，钢模面板，大长条零件")
        print("     - 正则宽松（兼容老格式）")
        print("     - 搜索半径大（最大5000mm），覆盖远处文字")
        print("     - 序号共享：排版图同序号多零件可共用")
        print()

        try:
            choice = input("请选择文字提取模式[1-2，默认1]: ").strip() or "1"
        except Exception as e:
            logger.warning(f"输入异常，使用默认: {e}")
            choice = "1"

        if choice == '2':
            self.text_mode = 'loose'
            self.text_extractor.text_mode = 'loose'
            print("   文字模式: 宽松覆盖（适合钢模面板/大长条）")
        else:
            self.text_mode = 'compact'
            self.text_extractor.text_mode = 'compact'
            print("   文字模式: 紧凑精确（适合规范制图）")
        print('='*60)

        # ===== 外框剥离模式选择 =====
        print("\n【外框剥离模式】")
        print("  0. 【不剥离】不识别外框，全部当作零件处理")
        print("  1. 【保守模式】v5.37z4 逻辑，安全不易拆散零件（默认）")
        print("  2. 【精准模式】v5.37j 逻辑，能剥钢模/排版图外框")
        print("  3. 【精准+自定义】精准模式 + 可调整最小外框阈值")
        print()
        try:
            frame_choice = input("请选择外框模式[0-3，默认1]: ").strip() or "1"
        except Exception:
            frame_choice = "1"

        if frame_choice == '0':
            self.frame_mode = 'none'
            print("   外框模式: 不剥离")
        elif frame_choice == '2':
            self.frame_mode = 'precise'
            self.frame_threshold = 300
            print("   外框模式: 精准模式 (阈值300mm)")
        elif frame_choice == '3':
            self.frame_mode = 'precise'
            try:
                self.frame_threshold = float(input("最小外框对角线阈值mm[300]: ") or "300")
            except Exception:
                self.frame_threshold = 300
            print(f"   外框模式: 精准+自定义 (阈值{self.frame_threshold}mm)")
        else:
            self.frame_mode = 'conservative'
            print("   外框模式: 保守模式")
        print('='*60)

    def process_file(self, input_path: str, output_dir: str) -> int:
        input_path = Path(input_path)
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # 清空旧零件文件，确保同路径重复拆分结果干净
        for old_dxf in output_dir.glob('**/*.dxf'):
            try:
                old_dxf.unlink()
            except Exception:
                pass

        self.drawing_no, self.piece_count, self.full_prefix = self._extract_drawing_no(input_path)
        self.frame_entities = []

        print(f"\n{'='*60}")
        print(f"图号: {self.drawing_no} | 件数: {self.piece_count}")
        print(f"前缀: {self.full_prefix}")
        print(f"处理: {input_path.name}")
        print(f"文字模式: {self.text_mode}")
        print(f"外框模式: {self.frame_mode}")
        print(f"{'='*60}")

        try:
            source_doc, auditor = recover.readfile(str(input_path))
            source_msp = source_doc.modelspace()

            print(f"\n  【步骤1】提取零件信息...")
            attr_texts, index_texts, process_texts, material_texts, split_texts, qty_texts = self.text_extractor.extract_all_texts(source_msp)
            print(f"  属性文字: {len(attr_texts)}个, 序号文字: {len(index_texts)}个, 材质标注: {len(material_texts)}个, 加工标注: {len(process_texts)}个, 拼接标记: {len(split_texts)}个, 独立数量: {len(qty_texts)}个")


            print(f"\n  【步骤2】提取几何实体...")
            entities = self._extract_all_entities(source_msp, source_doc)
            print(f"  总实体数: {len(entities)}")

            if len(entities) < 1:
                print(f"  ⚠ 实体过少")
                return 0

            print(f"\n  【步骤3】连通性分组...")
            connected_groups = self._build_connection_groups(entities)
            print(f"  连通性分组: {len(connected_groups)}个组")

            print(f"  【步骤4】外框检测...")
            if self.frame_mode == 'none':
                frame_groups, non_frame_groups = [], connected_groups
                print(f"  外框模式=不剥离，跳过检测")
            elif self.frame_mode == 'precise':
                frame_groups, non_frame_groups = self._detect_frames_by_nesting_precise(connected_groups, attr_texts)
                print(f"  外框组: {len(frame_groups)}个, 非外框组: {len(non_frame_groups)}个 (精准模式)")
            else:
                frame_groups, non_frame_groups = self._detect_frames_by_nesting(connected_groups, attr_texts)
                print(f"  外框组: {len(frame_groups)}个, 非外框组: {len(non_frame_groups)}个 (保守模式)")

            print(f"  【步骤4.5】包含聚合(孔合并)...")
            non_frame_groups = self._merge_holes_into_parts(non_frame_groups)
            print(f"  合并后非外框组: {len(non_frame_groups)}个")

            print(f"  【步骤5】邻近聚合(非外框组)...")
            parts = self._cluster_by_proximity(non_frame_groups)
            valid_parts = [p for p in parts if self._get_diagonal_size(p['bounds']) >= self.min_part_size]
            print(f"  有效零件: {len(valid_parts)}个")

            print(f"\n  【步骤6】导出并匹配命名...")
            success_count = self._export_and_match(
                source_doc, valid_parts, attr_texts, index_texts, process_texts, material_texts, split_texts, qty_texts, output_dir
            )

            if self.frame_entities:
                self._export_frames(source_doc, output_dir)
                self.stats['frames'] += len(self.frame_entities)

            self._generate_batch_import_excel(output_dir)

            self.stats['files'] += 1
            return success_count

        except Exception as e:
            logger.error(f"处理文件异常: {e}")
            logger.error(traceback.format_exc())
            print(f"  ✗ 错误: {e}")
            return 0

    def _has_attr_nearby(self, bounds, attr_texts, threshold_ratio=0.3, min_threshold=1500):
        if not attr_texts:
            return False
        diag = self._get_diagonal_size(bounds)
        threshold = max(diag * threshold_ratio, min_threshold)
        center_x = (bounds[0] + bounds[2]) / 2
        center_y = (bounds[1] + bounds[3]) / 2
        for attr in attr_texts:
            dist = math.hypot(attr['pos'][0] - center_x, attr['pos'][1] - center_y)
            if dist < threshold:
                return True
        return False

    def _detect_frames_by_nesting(self, connected_groups, attr_texts=None):
        if not connected_groups:
            return [], []

        group_infos = []
        for group in connected_groups:
            if not group:
                continue
            bounds = self._compute_bounds(group)
            diag = self._get_diagonal_size(bounds)
            group_infos.append({
                'group': group,
                'bounds': bounds,
                'diag': diag,
                'is_part': False,
                'children': [],
            })

        n = len(group_infos)
        if n == 0:
            return [], []

        for i in range(n):
            for j in range(n):
                if i != j:
                    bi = group_infos[i]['bounds']
                    bj = group_infos[j]['bounds']
                    if (bi[0] <= bj[0] and bi[1] <= bj[1] and 
                        bi[2] >= bj[2] and bi[3] >= bj[3]):
                        i_area = (bi[2]-bi[0]) * (bi[3]-bi[1])
                        j_area = (bj[2]-bj[0]) * (bj[3]-bj[1])
                        if j_area < i_area * 0.5 and j_area > 0:
                            group_infos[i]['children'].append(j)

        for gi in group_infos:
            gi['is_part'] = self._is_independent_part(gi['group'], gi['bounds'])

        big_groups = [i for i, gi in enumerate(group_infos) if gi['diag'] > 300]
        if len(big_groups) == 1:
            big_idx = big_groups[0]
            big_group = group_infos[big_idx]['group']
            big_bounds = group_infos[big_idx]['bounds']
            rect_check = self._is_rectangular_frame(big_group, big_bounds)

            if not rect_check['is_rect'] and not self._has_attr_nearby(big_bounds, attr_texts):
                children = group_infos[big_idx]['children']
                if len(children) >= 2:
                    hole_like = 0
                    part_like = 0
                    for c in children:
                        if self._is_likely_hole(big_group, group_infos[c]['group']):
                            hole_like += 1
                        else:
                            part_like += 1

                    if hole_like > part_like:
                        return [], [gi['group'] for gi in group_infos]

        frame_groups = []
        non_frame_groups = []

        for i, gi in enumerate(group_infos):
            group = gi['group']
            bounds = gi['bounds']
            diag = gi['diag']
            children = gi['children']

            rect_check = self._is_rectangular_frame(group, bounds)
            
            # 【v5.37o 核心修复】如果组内包含ARC/ELLIPSE/SPLINE曲线，强制不视为矩形外框
            # 避免弧形零件两端的直线平台被误判为矩形
            has_curves = any(e['type'] in ('ARC', 'ELLIPSE', 'SPLINE') for e in group)
            if has_curves:
                rect_check['is_rect'] = False
            
            is_closed = rect_check['is_rect'] or self._is_closed_loop(group)

            is_frame = False
            if is_closed and diag > 300:
                entity_count = len(group)
                part_children = [c for c in children 
                                 if not self._is_likely_hole(group, group_infos[c]['group'])]
                is_simple_outline = rect_check['is_rect'] or entity_count <= 16

                if rect_check['is_rect'] and len(children) >= 2 and len(part_children) >= 1:
                    is_frame = True
                # 非矩形轮廓即使有零件子组，也不再剥外框（避免拆散弧形/异形零件）

            if is_frame:
                frame_groups.append(group)
                for e in group:
                    self.frame_entities.append(e)
            else:
                non_frame_groups.append(group)

        return frame_groups, non_frame_groups


    # ==================== 精准外框检测 (v5.37j 完整复制) ====================
    def _detect_frames_by_nesting_precise(self, connected_groups, attr_texts=None):
        if not connected_groups:
            return [], []

        group_infos = []
        for group in connected_groups:
            if not group:
                continue
            bounds = self._compute_bounds(group)
            diag = self._get_diagonal_size(bounds)
            group_infos.append({
                'group': group,
                'bounds': bounds,
                'diag': diag,
                'is_part': False,
                'children': [],
            })

        n = len(group_infos)
        if n == 0:
            return [], []

        for i in range(n):
            for j in range(n):
                if i != j:
                    bi = group_infos[i]['bounds']
                    bj = group_infos[j]['bounds']
                    if (bi[0] <= bj[0] and bi[1] <= bj[1] and 
                        bi[2] >= bj[2] and bi[3] >= bj[3]):
                        i_area = (bi[2]-bi[0]) * (bi[3]-bi[1])
                        j_area = (bj[2]-bj[0]) * (bj[3]-bj[1])
                        if j_area < i_area * 0.5 and j_area > 0:
                            group_infos[i]['children'].append(j)

        for gi in group_infos:
            gi['is_part'] = self._is_independent_part(gi['group'], gi['bounds'])

        big_groups = [i for i, gi in enumerate(group_infos) if gi['diag'] > self.frame_threshold]
        if len(big_groups) == 1:
            big_idx = big_groups[0]
            big_group = group_infos[big_idx]['group']
            big_bounds = group_infos[big_idx]['bounds']
            rect_check = self._is_rectangular_frame(big_group, big_bounds)

            if not rect_check['is_rect'] and not self._has_attr_nearby(big_bounds, attr_texts):
                children = group_infos[big_idx]['children']
                if len(children) >= 2:
                    hole_like = 0
                    part_like = 0
                    for c in children:
                        if group_infos[c]['is_part']:
                            part_like += 1
                        else:
                            hole_like += 1

                    if hole_like > part_like:
                        return [], [gi['group'] for gi in group_infos]

        frame_groups = []
        non_frame_groups = []

        for i, gi in enumerate(group_infos):
            group = gi['group']
            bounds = gi['bounds']
            diag = gi['diag']
            children = gi['children']

            rect_check = self._is_rectangular_frame(group, bounds)
            is_closed = rect_check['is_rect'] or self._is_closed_loop(group)

            is_frame = False
            if is_closed and diag > self.frame_threshold:
                entity_count = len(group)
                part_children = [c for c in children if group_infos[c]['is_part']]
                is_simple_outline = rect_check['is_rect'] or entity_count <= 16

                if rect_check['is_rect'] and len(children) >= 2 and len(part_children) >= 1:
                    is_frame = True
                elif (len(part_children) >= 2 and is_simple_outline 
                      and not self._has_attr_nearby(bounds, attr_texts)):
                    is_frame = True

            if is_frame:
                frame_groups.append(group)
                for e in group:
                    self.frame_entities.append(e)
            else:
                non_frame_groups.append(group)

        return frame_groups, non_frame_groups

    def _is_independent_part(self, group, bounds):
        diag = self._get_diagonal_size(bounds)

        if diag < self.min_part_size * 2:
            return False

        rect_check = self._is_rectangular_frame(group, bounds)
        is_closed = rect_check['is_rect'] or self._is_closed_loop(group)
        if not is_closed:
            return False

        if all(e['type'] in ['CIRCLE', 'ARC'] for e in group):
            return False

        if (len(group) == 1 and group[0]['type'] == 'LWPOLYLINE' 
            and rect_check['is_rect']):
            return False

        if self._is_line_rect(group, bounds) and diag < 150:
            return False

        if diag < 100 and len(group) <= 12:
            simple_types = {'LINE', 'ARC', 'CIRCLE', 'LWPOLYLINE'}
            if all(e['type'] in simple_types for e in group):
                return False

        if is_closed and diag < 200:
            w = bounds[2] - bounds[0]
            h = bounds[3] - bounds[1]
            if min(w, h) > 0 and max(w, h) / min(w, h) > 5:
                return False

        return True

    def _is_likely_hole(self, parent_entities, child_entities):
        """判断子组是否为孔（被包含的小闭合轮廓），而非独立零件"""
        parent_bbox = self._compute_bounds(parent_entities)
        child_bbox = self._compute_bounds(child_entities)
        
        p_w = parent_bbox[2] - parent_bbox[0]
        p_h = parent_bbox[3] - parent_bbox[1]
        c_w = child_bbox[2] - child_bbox[0]
        c_h = child_bbox[3] - child_bbox[1]
        
        p_area = max(p_w * p_h, 1)
        c_area = max(c_w * c_h, 1)
        
        area_ratio = c_area / p_area
        child_diag = math.sqrt(c_w**2 + c_h**2)
        
        if area_ratio < 0.15:
            return True
        
        if all(e['type'] == 'CIRCLE' for e in child_entities) and area_ratio < 0.25:
            return True
        
        return area_ratio < 0.25 and child_diag < 300

    def _merge_holes_into_parts(self, groups):
        """包含聚合：把被包含的小组合并到父组（孔合并回零件）"""
        if len(groups) < 2:
            return groups
        
        n = len(groups)
        merged = [False] * n
        
        infos = []
        for g in groups:
            b = self._compute_bounds(g)
            area = max((b[2]-b[0]) * (b[3]-b[1]), 1)
            infos.append({'group': g, 'bounds': b, 'area': area})
        
        order = sorted(range(n), key=lambda i: infos[i]['area'], reverse=True)
        
        for pi in order:
            if merged[pi]:
                continue
            p_bounds = infos[pi]['bounds']
            p_area = infos[pi]['area']
            
            for ci in range(n):
                if ci == pi or merged[ci]:
                    continue
                c_bounds = infos[ci]['bounds']
                c_area = infos[ci]['area']
                
                if (c_bounds[0] >= p_bounds[0] and c_bounds[1] >= p_bounds[1] and
                    c_bounds[2] <= p_bounds[2] and c_bounds[3] <= p_bounds[3]):
                    
                    if c_area / p_area < 0.12:
                        infos[pi]['group'].extend(infos[ci]['group'])
                        merged[ci] = True
        
        return [infos[i]['group'] for i in range(n) if not merged[i]]

    def _is_line_rect(self, group, bounds):
        if len(group) != 4:
            return False
        if not all(e['type'] == 'LINE' for e in group):
            return False
        tol = self.gap_tolerance * 2
        h_lines = []
        v_lines = []
        min_x, min_y, max_x, max_y = bounds
        for e in group:
            s = e['entity'].dxf.start
            end = e['entity'].dxf.end
            dx = abs(end[0] - s[0])
            dy = abs(end[1] - s[1])
            if dy < tol and dx > tol:
                h_lines.append((s[1] + end[1]) / 2)
            elif dx < tol and dy > tol:
                v_lines.append((s[0] + end[0]) / 2)
        if len(h_lines) != 2 or len(v_lines) != 2:
            return False
        has_top = any(abs(y - max_y) < tol * 5 for y in h_lines)
        has_bottom = any(abs(y - min_y) < tol * 5 for y in h_lines)
        has_left = any(abs(x - min_x) < tol * 5 for x in v_lines)
        has_right = any(abs(x - max_x) < tol * 5 for x in v_lines)
        return has_top and has_bottom and has_left and has_right

    def _is_closed_loop(self, group):
        if not group:
            return False

        endpoints = []
        for e in group:
            eps = e.get('endpoints', [])
            endpoints.extend(eps)

        if len(endpoints) < 2:
            return False

        tol = self.gap_tolerance * 2
        rounded = [(round(p[0]/tol), round(p[1]/tol)) for p in endpoints]
        counts = Counter(rounded)

        single_count = sum(1 for c in counts.values() if c == 1)
        return single_count <= 2

    def _is_rectangular_frame(self, entities: List[Dict], bounds: Tuple[float, float, float, float]) -> Dict:
        min_x, min_y, max_x, max_y = bounds
        tol = self.gap_tolerance * 2
        lines = []
        lwpolylines = []
        for e in entities:
            if e['type'] == 'LINE':
                s = e['entity'].dxf.start
                end = e['entity'].dxf.end
                lines.append({'x1': s[0], 'y1': s[1], 'x2': end[0], 'y2': end[1]})
            elif e['type'] == 'LWPOLYLINE':
                lwpolylines.append(e)
        for lwp in lwpolylines:
            try:
                pts = list(lwp['entity'].vertices_in_wcs())
                n_pts = len(pts)
                if n_pts >= 4 and n_pts <= 6:
                    xs = [p[0] for p in pts[:min(n_pts, 5)]]
                    ys = [p[1] for p in pts[:min(n_pts, 5)]]
                    unique_x = sorted(set(round(x/tol)*tol for x in xs))
                    unique_y = sorted(set(round(y/tol)*tol for y in ys))
                    if len(unique_x) == 2 and len(unique_y) == 2:
                        return {'is_rect': True, 'lines': [], 'rect_bounds': bounds, 
                                'h_lines': [], 'v_lines': [], 'is_lwpolyline': True, 'lwpolyline': lwp}
            except Exception:
                pass
        if len(lines) < 4:
            return {'is_rect': False, 'lines': [], 'rect_bounds': bounds, 
                    'h_lines': [], 'v_lines': [], 'is_lwpolyline': False}
        h_lines = []
        v_lines = []
        for line in lines:
            dx = abs(line['x2'] - line['x1'])
            dy = abs(line['y2'] - line['y1'])
            if dy < tol and dx > tol:
                h_lines.append(line)
            elif dx < tol and dy > tol:
                v_lines.append(line)
        if len(h_lines) < 2 or len(v_lines) < 2:
            return {'is_rect': False, 'lines': lines, 'rect_bounds': bounds, 
                    'h_lines': h_lines, 'v_lines': v_lines, 'is_lwpolyline': False}
        h_y_values = [(line['y1'] + line['y2']) / 2 for line in h_lines]
        v_x_values = [(line['x1'] + line['x2']) / 2 for line in v_lines]
        has_top = any(abs(y - max_y) < tol * 5 for y in h_y_values)
        has_bottom = any(abs(y - min_y) < tol * 5 for y in h_y_values)
        has_left = any(abs(x - min_x) < tol * 5 for x in v_x_values)
        has_right = any(abs(x - max_x) < tol * 5 for x in v_x_values)
        is_rect = has_top and has_bottom and has_left and has_right
        return {'is_rect': is_rect, 'lines': lines, 'rect_bounds': bounds, 
                'h_lines': h_lines, 'v_lines': v_lines, 'is_lwpolyline': False}

    def _compute_bounds(self, entities: List[Dict]) -> Tuple[float, float, float, float]:
        if not entities:
            return (0, 0, 0, 0)
        all_bounds = [e['bounds'] for e in entities]
        return (min(b[0] for b in all_bounds), min(b[1] for b in all_bounds),
                max(b[2] for b in all_bounds), max(b[3] for b in all_bounds))

    def _export_frames(self, source_doc, output_dir: Path):
        if not self.frame_entities:
            return
        frame_dir = output_dir / "外框"
        frame_dir.mkdir(parents=True, exist_ok=True)
        try:
            target_doc = ezdxf.new('R2010')
            target_msp = target_doc.modelspace()
            importer = Importer(source_doc, target_doc)
            all_frame_ents = []
            for e in self.frame_entities:
                if isinstance(e, dict) and 'entity' in e:
                    all_frame_ents.append(e['entity'])
                elif hasattr(e, 'dxftype'):
                    all_frame_ents.append(e)
            if all_frame_ents:
                importer.import_entities(all_frame_ents, target_msp)
                importer.finalize()
                frame_path = frame_dir / f"{self.full_prefix}_外框.dxf"
                target_doc.saveas(str(frame_path))
                print(f"    外框已导出: {frame_path.name} ({len(all_frame_ents)}个实体)")
        except Exception as e:
            logger.warning(f"外框导出失败: {e}")
    def _find_material_compact(self, search_center, material_texts, part_diag, used_materials=None):
        if not material_texts or not search_center:
            return 'Q235'
        # [v5.37z6 修复] 材质搜索半径放宽，钢模面板材质位置不确定
        # [v5.37z6-修复] 缩小搜索半径，防止远处材质被抢，宁缺勿乱
        # [v5.37z6-修复] 分级搜索半径：小零件不抢，大钢模能覆盖远处
        best = None
        best_dist = float('inf')
        for idx, mt in enumerate(material_texts):
            # [v5.37z6-修复] 恢复 used_materials 独占，一个材质标注只配一个零件
            if used_materials is not None and idx in used_materials:
                continue
            dist = math.hypot(mt['pos'][0] - search_center[0], mt['pos'][1] - search_center[1])
            # [v5.37z6-修复] 距离超过对角线0.5倍时放弃，防止远处材质被抢
            if dist > part_diag * 0.5 + 100:
                continue
            if dist < best_dist:
                best_dist = dist
                best = idx
        if best is not None:
            # [v5.37z6-修复] 恢复独占：标记已用
            if used_materials is not None:
                used_materials.add(best)
            return material_texts[best]['material']
        return 'Q235'


    def _find_material_loose(self, search_center, material_texts, part_diag, used_materials=None, attr_pos=None):
        if not material_texts or not search_center:
            return 'Q235'
        # [松散模式] 材质搜索半径更大，覆盖远处材质
        # [v5.37z17-修复] 如果已有属性文字，小零件收紧防抢，大零件保持原样
        if attr_pos is not None:
            search_center = attr_pos
            if part_diag < 500:
                # 小零件严格：半径500mm，防止抢远处其他零件的材质
                max_search = 500
            else:
                # 大零件保持原样：宽松半径，钢模/细长/法兰不受影响
                max_search = min(max(part_diag * 0.5 + 500, 800), 3000)
        else:
            max_search = min(max(part_diag * 0.5 + 500, 800), 3000)
        best = None
        best_dist = float('inf')
        for idx, mt in enumerate(material_texts):
            if used_materials is not None and idx in used_materials:
                continue
            dist = math.hypot(mt['pos'][0] - search_center[0], mt['pos'][1] - search_center[1])
            if dist > max_search:
                continue
            if dist < best_dist:
                best_dist = dist
                best = idx
        if best is not None:
            # [v5.37z6-修复] 恢复独占：标记已用
            if used_materials is not None:
                used_materials.add(best)
            return material_texts[best]['material']
        return 'Q235'


    def _find_material(self, search_center, material_texts, part_diag, used_materials=None, attr_pos=None):
        if self.text_mode == 'loose':
            return self._find_material_loose(search_center, material_texts, part_diag, used_materials, attr_pos)
        return self._find_material_compact(search_center, material_texts, part_diag, used_materials)


    def _mark_material_used(self, material_val, text_pos, material_texts, used_materials):
        if not material_val or not text_pos or not material_texts:
            return
        for midx, mt in enumerate(material_texts):
            if midx in used_materials:
                continue
            if mt['material'] == material_val:
                dist = math.hypot(mt['pos'][0] - text_pos[0], mt['pos'][1] - text_pos[1])
                if dist < 1.0:
                    used_materials.add(midx)
                    return

    def _export_and_match_compact(self, source_doc, parts, attr_texts, index_texts, process_texts, material_texts, split_texts, qty_texts, output_dir):
        """
        [融合版 v5.37v] 采用 v5.24.8 的贪心最近匹配算法，
        避免前面的零件抢走后面的属性文字，同时保留 v5.37t 的命名规范
        """
        success = 0
        thickness_dirs = {}
        other_dir = output_dir / "其他"
        other_dir.mkdir(parents=True, exist_ok=True)

        used_attrs = set()
        used_indices = set()
        used_process = set()

        used_materials = set()
        used_qtys = set()
        used_sub_indices = defaultdict(set)  # (index, thickness) -> set of used sub_indices


        # ========== 第一步：预计算所有零件到所有属性文字的距离 ==========
        print(f"\n  【贪心最近匹配】共{len(parts)}个零件，{len(attr_texts)}个属性文字")

        distance_pairs = []
        for i, part in enumerate(parts):
            part_center = part['center']
            for j, attr in enumerate(attr_texts):
                dist = math.hypot(part_center[0] - attr['pos'][0], part_center[1] - attr['pos'][1])
                distance_pairs.append({
                    'part_idx': i,
                    'attr_idx': j,
                    'distance': dist,
                })

        distance_pairs.sort(key=lambda x: x['distance'])

        # ========== 第二步：贪心匹配 - 每个属性文字匹配给最近的未匹配零件 ==========
        matched_parts = {}  # part_idx -> attr_idx


        for pair in distance_pairs:
            part_idx = pair['part_idx']
            attr_idx = pair['attr_idx']
            dist = pair['distance']

            # [v5.37z6-修复15] 动态最大距离：小零件500mm，中等800mm，大零件1500mm
            pb = parts[part_idx]['bounds']
            part_diag = math.hypot(pb[2]-pb[0], pb[3]-pb[1])
            if part_diag < 300:
                max_dist = 500
            elif part_diag < 1500:
                max_dist = 800
            else:
                max_dist = 1500

            if part_idx in matched_parts or dist > max_dist:
                continue

            matched_parts[part_idx] = attr_idx

        print(f"    成功匹配: {len(matched_parts)}对")

        # ========== 第二步b：全局贪心匹配加工标注 ==========
        print(f"  【加工标注全局匹配】共{len(parts)}个零件，{len(process_texts)}个加工标注")
        proc_pairs = []
        for i, part in enumerate(parts):
            part_center = part['center']
            pb = part['bounds']
            part_diag = self._get_diagonal_size(pb)
            if part_diag < 300:
                proc_max = 400
            elif part_diag < 1500:
                proc_max = 1000
            else:
                proc_max = 2000
            for j, pt in enumerate(process_texts):
                dist = math.hypot(part_center[0] - pt['pos'][0], part_center[1] - pt['pos'][1])
                if dist <= proc_max:
                    proc_pairs.append({'part_idx': i, 'proc_idx': j, 'distance': dist})
        proc_pairs.sort(key=lambda x: x['distance'])

        matched_proc = defaultdict(list)
        used_proc_indices = set()
        for pair in proc_pairs:
            if pair['proc_idx'] in used_proc_indices:
                continue
            matched_proc[pair['part_idx']].append(pair['proc_idx'])
            used_proc_indices.add(pair['proc_idx'])
        print(f"    成功匹配: {sum(len(v) for v in matched_proc.values())}对")

        # ========== 第三步：按匹配结果导出零件 ==========
        for i, part in enumerate(parts, 1):
            temp_result = self._export_part_temp(source_doc, part, f"_temp_{i:04d}.dxf", output_dir)
            if not temp_result:
                continue

            temp_path = temp_result['path']
            part_center = temp_result['center']
            part_bounds = temp_result['bounds']
            part_diag = self._get_diagonal_size(part_bounds)

            # [v5.37z6-修复] 分级搜索半径
            split_search_radius = 500  # [v5.37z6-修复] 固定500mm，拼接标记通常就在零件旁边
            # [v5.37z6 修复] 加工搜索半径按长边放宽，长条零件加工标注在上下方
            w = part_bounds[2] - part_bounds[0]
            h = part_bounds[3] - part_bounds[1]
            # [v5.37z6-修复] 缩小搜索半径，防止远处加工标注被抢
            # [v5.37z6-修复] 分级搜索半径
            if part_diag < 300:
                proc_search_radius = 400
            elif part_diag < 1500:
                proc_search_radius = 1000
            else:
                proc_search_radius = 2000

            # 检查是否有匹配的属性文字
            if (i-1) in matched_parts:
                attr_idx = matched_parts[i-1]
                best_attr = attr_texts[attr_idx]
                used_attrs.add(attr_idx)

                thickness = best_attr['thickness']
                index_val = best_attr['index']
                qty = best_attr['qty']
                has_qty = best_attr['has_qty']

                # [v5.37z6-修复3] 如果属性文字只有厚度没有数量（如T6），到附近找独立数量标注（如*4）
                if not has_qty and qty_texts:
                    best_qty = None
                    best_qty_dist = float('inf')
                    for qt in qty_texts:
                        if id(qt) in used_qtys:
                            continue
                        dist = math.hypot(qt['pos'][0] - part_center[0], qt['pos'][1] - part_center[1])
                        # 独立数量标注通常就在厚度旁边，距离很近
                        if dist < 200 and dist < best_qty_dist:
                            best_qty_dist = dist
                            best_qty = qt
                    if best_qty:
                        qty = best_qty['qty']
                        has_qty = True
                        used_qtys.add(id(best_qty))

                # [v5.37z4 修复] 恢复序号补充：有T标注的零件可从附近序号文字补充序号
                # （T标注和序号通常是两个独立实体，T标注本身不含序号）
                matched_idx_text = None
                if index_val == 'xx':
                    best_idx = None
                    best_idx_dist = float('inf')
                    # [v5.37z6 修复] 动态搜索半径：小零件封顶800，大零件放开5000
                    idx_search_radius = max(part_diag * 1.5, 300)
                    if part_diag > 1500:
                        idx_search_radius = 5000
                    else:
                        idx_search_radius = min(idx_search_radius, 800)
                    for idx, idx_text in enumerate(index_texts):
                        if idx in used_indices:
                            continue
                        dist = math.hypot(idx_text['pos'][0] - part_center[0],
                                        idx_text['pos'][1] - part_center[1])
                        if dist < idx_search_radius and dist < best_idx_dist:
                            best_idx_dist = dist
                            best_idx = idx
                            index_val = idx_text['index']
                            matched_idx_text = idx_text

                    if best_idx is not None:
                        # [v5.37z8-保险] 小零件序号距离超过阈值时放弃
                        # 默认阈值: part_diag * 0.8，可通过临时参数覆盖
                        idx_max = self.config.get('idx_search_max', part_diag * 0.8)
                        if part_diag < 500 and best_idx_dist > idx_max:
                            best_idx = None
                            index_val = 'xx'
                            matched_idx_text = None
                        else:
                            # [v5.37z6-修复] 恢复独占：标记已用
                            used_indices.add(best_idx)

                # 加工码
                process_code_chars = ['x'] * 5
                attr_proc = self.text_extractor.parse_process_code(best_attr['content'])
                for j in range(5):
                    if attr_proc[j] != 'x':
                        process_code_chars[j] = attr_proc[j]

                if matched_idx_text:
                    idx_proc = self.text_extractor.parse_process_code(matched_idx_text['content'])
                    for j in range(5):
                        if idx_proc[j] != 'x' and process_code_chars[j] == 'x':
                            process_code_chars[j] = idx_proc[j]

                                # 使用全局匹配的独立加工标注（允许多个）
                if (i-1) in matched_proc:
                    for proc_idx in matched_proc[i-1]:
                        pt = process_texts[proc_idx]
                        for j in range(5):
                            if pt['process_code'][j] != 'x' and process_code_chars[j] == 'x':
                                process_code_chars[j] = pt['process_code'][j]
                        used_process.add(proc_idx)

                process_code = ''.join(process_code_chars)

                # 材质
                material = 'Q235'
                attr_mat = self.text_extractor.parse_material(best_attr.get('content', ''))
                if attr_mat:
                    material = attr_mat
                    self._mark_material_used(material, best_attr['pos'], material_texts, used_materials)

                if material == 'Q235' and matched_idx_text:
                    # [v5.37z 修复2] 序号文字带#号，不应从中提取材质
                    idx_content = matched_idx_text.get('content', '')
                    if '#' not in idx_content and '＃' not in idx_content:
                        idx_mat = self.text_extractor.parse_material(idx_content)
                        if idx_mat:
                            material = idx_mat
                            self._mark_material_used(material, matched_idx_text['pos'], material_texts, used_materials)

                if material == 'Q235':
                    material = self._find_material(part_center, material_texts, part_diag, used_materials)

                # 拼接标记
                sub_index = None
                is_original_sub = False  # [v5.37z8-修复] 标记来源：True=图纸原生, False=自动递增

                # [v5.37z6-修复] 先匹配 split_texts（如7#-2 → _02）
                if index_val != 'xx':
                    candidates = []
                    for idx, st in enumerate(split_texts):
                        if st.get('index') == index_val:
                            dist = math.hypot(st['pos'][0] - part_center[0], st['pos'][1] - part_center[1])
                            if dist < split_search_radius:
                                candidates.append((dist, idx, st['sub_index']))
                    if candidates:
                        candidates.sort()
                        _, best_idx, sub_index = candidates[0]
                        is_original_sub = True  # 来自图纸原生标注
                        # 有拼接标记，不递增，保持原样

                # [v5.37z6-修复] 再匹配无index的 split_texts
                if sub_index is None:
                    candidates = []
                    for idx, st in enumerate(split_texts):
                        if st.get('index') is None:
                            dist = math.hypot(st['pos'][0] - part_center[0], st['pos'][1] - part_center[1])
                            if dist < split_search_radius:
                                candidates.append((dist, idx, st['sub_index']))
                    if candidates:
                        candidates.sort()
                        _, best_idx, sub_index = candidates[0]
                        is_original_sub = True  # 来自图纸原生标注

                # [v5.37z6-修复2] 如果 split_texts 没提取到，直接从匹配到的属性文字中解析拼接标记
                if sub_index is None and index_val != 'xx' and best_attr:
                    attr_content = best_attr.get('content', '')
                    if attr_content:
                        # 从属性文字中直接提取拼接标记，如 7#-2 → _02, 6#_3 → _03
                        direct_split = re.search(r'(?<![A-Z0-9])' + re.escape(index_val) + r'[#＃]-(\d+)(?![0-9#＃])', attr_content.upper())
                        if not direct_split:
                            direct_split = re.search(r'(?<![A-Z0-9])' + re.escape(index_val) + r'[#＃]_(\d+)(?![0-9#＃])', attr_content.upper())
                        if direct_split:
                            sub_index = f"{int(direct_split.group(1)):02d}"
                            is_original_sub = True  # 来自图纸原生标注

                # [v5.37z6-修复] 没有拼接标记时，同序号同厚度自动递增
                if sub_index is None and index_val != 'xx' and thickness > 0:
                    combo_key = (index_val, thickness)
                    candidate = 1
                    while f"{candidate:02d}" in used_sub_indices.get(combo_key, set()):
                        candidate += 1
                    sub_index = f"{candidate:02d}"
                    used_sub_indices[combo_key].add(sub_index)
                    is_original_sub = False  # 自动递增生成

                # [v5.37z6-修复] xx# 零件递增
                if sub_index is None and index_val == 'xx' and thickness > 0:
                    xx_key = (index_val, thickness)
                    xx_idx = 1
                    while f"{xx_idx:02d}" in used_sub_indices.get(xx_key, set()):
                        xx_idx += 1
                    sub_index = f"{xx_idx:02d}"
                    used_sub_indices[xx_key].add(sub_index)
                    is_original_sub = False  # 自动递增生成

                if sub_index:
                    index_display = f"{index_val}#_{sub_index}"
                else:
                    index_display = f"{index_val}#"

                # [v5.37z6-修复] 补全缺失的 thickness_str、qty_str、target_dir
                thickness_str = f"T{thickness}"
                qty_str = str(qty)
                # [v5.37z8-文件夹结构] 材质/表面状态/厚度 三级结构，有则建无则跳过
                surface = self.text_extractor.parse_surface(best_attr.get('content', ''))
                if not surface and matched_idx_text:
                    surface = self.text_extractor.parse_surface(matched_idx_text.get('content', ''))

                material_str = material  # 如 Q235, 304
                dir_key = (material_str, thickness_str, surface)
                target_dir = thickness_dirs.get(dir_key)
                if not target_dir:
                    material_dir = output_dir / material_str
                    material_dir.mkdir(parents=True, exist_ok=True)
                    if surface:
                        surface_dir = material_dir / surface
                        surface_dir.mkdir(parents=True, exist_ok=True)
                        target_dir = surface_dir / thickness_str
                    else:
                        target_dir = material_dir / thickness_str
                    target_dir.mkdir(parents=True, exist_ok=True)
                    thickness_dirs[dir_key] = target_dir

                new_name = f"{self.full_prefix}_{index_display}_{material}_{thickness_str}_B_{process_code}_{qty_str}.dxf"
                print(f"  ✓ {thickness_str}/{new_name} (材质{material})")
                success += 1
                self.stats['matched'] += 1

            else:
                # [v5.37z3 保留] 无属性分支：不借用任何序号和厚度，直接进"其他"
                index_val = 'xx'
                thickness = 0
                process_code = 'xxxxx'
                material = 'Q235'
                qty_str = "1"

                target_dir = other_dir
                thickness_str = "T未知"

                new_name = f"{self.full_prefix}_P{i:02d}_{material}_{thickness_str}_B_{process_code}_{qty_str}.dxf"
                print(f"  ⚠ 其他/{new_name} (无属性，不借用)")
                self.stats['unmatched'] += 1

            new_path = target_dir / new_name
            # [v5.37z8-修复] 区分原生 sub_index 和自动递增 sub_index 的冲突处理
            if new_path.exists() and index_val != 'xx' and thickness > 0:
                if is_original_sub:
                    # 图纸原生 sub_index（如14#-2）重复，直接扔"其他"，不递增
                    target_dir = other_dir
                    new_path = target_dir / new_name
                    print(f"  ⚠ 原生拼接标记重复，放其他: {new_name}")
                else:
                    # 自动递增 sub_index（如14#_01）重复，继续递增
                    base_sub = int(sub_index) if sub_index else 1
                    candidate = base_sub + 1
                    while candidate <= 99:
                        index_display = f"{index_val}#_{candidate:02d}"
                        new_name = f"{self.full_prefix}_{index_display}_{material}_{thickness_str}_B_{process_code}_{qty_str}.dxf"
                        new_path = target_dir / new_name
                        if not new_path.exists():
                            break
                        candidate += 1
                    if candidate <= 99:
                        print(f"  ⚠ 自动递增: {new_name}")
                    else:
                        target_dir = other_dir
                        new_path = target_dir / new_name
                        print(f"  ⚠ 递增到99仍冲突，放其他: {new_name}")
            if new_path.exists():
                print(f"  ⚠ 覆盖已存在文件: {new_name}")

            try:
                shutil.move(str(temp_path), str(new_path))
            except Exception as e:
                print(f"    ⚠ 移动失败: {e}")

        return success

    def _export_and_match_loose(self, source_doc, parts, attr_texts, index_texts, process_texts, material_texts, split_texts, qty_texts, output_dir):
        """
        [松散模式] 采用早期版 v5.24.8 宽松匹配逻辑 采用 v5.24.8 的贪心最近匹配算法，
        避免前面的零件抢走后面的属性文字，同时保留 v5.37t 的命名规范
        """
        success = 0
        thickness_dirs = {}
        other_dir = output_dir / "其他"
        other_dir.mkdir(parents=True, exist_ok=True)

        used_attrs = set()
        used_indices = set()
        used_process = set()
        used_splits = set()
        used_materials = set()
        used_sub_indices = defaultdict(set)  # (index, thickness) -> set of used sub_indices

        GLOBAL_MAX_DIST = 5000

        # ========== 第一步：预计算所有零件到所有属性文字的距离 ==========
        print(f"\n  【贪心最近匹配】共{len(parts)}个零件，{len(attr_texts)}个属性文字")

        distance_pairs = []
        for i, part in enumerate(parts):
            part_center = part['center']
            for j, attr in enumerate(attr_texts):
                dist = math.hypot(part_center[0] - attr['pos'][0], part_center[1] - attr['pos'][1])
                distance_pairs.append({
                    'part_idx': i,
                    'attr_idx': j,
                    'distance': dist,
                })

        distance_pairs.sort(key=lambda x: x['distance'])

        # ========== 第二步：贪心匹配 - 每个属性文字匹配给最近的未匹配零件 ==========
        matched_parts = {}  # part_idx -> attr_idx
        matched_attrs = set()

        for pair in distance_pairs:
            part_idx = pair['part_idx']
            attr_idx = pair['attr_idx']
            dist = pair['distance']

            if part_idx in matched_parts or attr_idx in matched_attrs or dist > GLOBAL_MAX_DIST:
                continue

            matched_parts[part_idx] = attr_idx
            matched_attrs.add(attr_idx)

        print(f"    成功匹配: {len(matched_parts)}对")

        # ========== 第二步b：全局贪心匹配加工标注 ==========
        print(f"  【加工标注全局匹配】共{len(parts)}个零件，{len(process_texts)}个加工标注")
        proc_pairs = []
        for i, part in enumerate(parts):
            part_center = part['center']
            pb = part['bounds']
            part_diag = self._get_diagonal_size(pb)
            # [松散模式] 恢复宽松半径，大钢模能覆盖远处加工标注
            proc_max = max(part_diag * 0.5 + 300, 500)
            for j, pt in enumerate(process_texts):
                dist = math.hypot(part_center[0] - pt['pos'][0], part_center[1] - pt['pos'][1])
                if dist <= proc_max:
                    proc_pairs.append({'part_idx': i, 'proc_idx': j, 'distance': dist})
        proc_pairs.sort(key=lambda x: x['distance'])

        matched_proc = defaultdict(list)
        used_proc_indices = set()
        for pair in proc_pairs:
            if pair['proc_idx'] in used_proc_indices:
                continue
            matched_proc[pair['part_idx']].append(pair['proc_idx'])
            used_proc_indices.add(pair['proc_idx'])
        print(f"    成功匹配: {sum(len(v) for v in matched_proc.values())}对")

        # ========== 第三步：按匹配结果导出零件 ==========
        for i, part in enumerate(parts, 1):
            temp_result = self._export_part_temp(source_doc, part, f"_temp_{i:04d}.dxf", output_dir)
            if not temp_result:
                continue

            temp_path = temp_result['path']
            part_center = temp_result['center']
            part_bounds = temp_result['bounds']
            part_diag = self._get_diagonal_size(part_bounds)

            split_search_radius = max(part_diag * 0.5 + 200, 300)
            proc_search_radius = max(part_diag * 0.5 + 300, 500)

            # 检查是否有匹配的属性文字
            if (i-1) in matched_parts:
                attr_idx = matched_parts[i-1]
                best_attr = attr_texts[attr_idx]
                used_attrs.add(attr_idx)

                thickness = best_attr['thickness']
                index_val = best_attr['index']
                qty = best_attr['qty']
                has_qty = best_attr['has_qty']

                # [v5.37z4 修复] 恢复序号补充：有T标注的零件可从附近序号文字补充序号
                # （T标注和序号通常是两个独立实体，T标注本身不含序号）
                matched_idx_text = None
                if index_val == 'xx':
                    best_idx = None
                    best_idx_dist = float('inf')
                    for idx, idx_text in enumerate(index_texts):
                        dist = math.hypot(idx_text['pos'][0] - part_center[0],
                                        idx_text['pos'][1] - part_center[1])
                        if dist < 5000 and dist < best_idx_dist:
                            best_idx_dist = dist
                            best_idx = idx
                            index_val = idx_text['index']
                            matched_idx_text = idx_text

                    if best_idx is not None:
                        # 不加入used_indices，允许多个零件共享同序号（排版图常见）
                        pass

                # 加工码
                process_code_chars = ['x'] * 5
                attr_proc = self.text_extractor.parse_process_code(best_attr['content'])
                for j in range(5):
                    if attr_proc[j] != 'x':
                        process_code_chars[j] = attr_proc[j]

                if matched_idx_text:
                    idx_proc = self.text_extractor.parse_process_code(matched_idx_text['content'])
                    for j in range(5):
                        if idx_proc[j] != 'x' and process_code_chars[j] == 'x':
                            process_code_chars[j] = idx_proc[j]

                                # 使用全局匹配的独立加工标注（允许多个）
                if (i-1) in matched_proc:
                    for proc_idx in matched_proc[i-1]:
                        pt = process_texts[proc_idx]
                        for j in range(5):
                            if pt['process_code'][j] != 'x' and process_code_chars[j] == 'x':
                                process_code_chars[j] = pt['process_code'][j]
                        used_process.add(proc_idx)

                process_code = ''.join(process_code_chars)

                # 材质
                material = 'Q235'
                attr_mat = self.text_extractor.parse_material(best_attr.get('content', ''))
                if attr_mat:
                    material = attr_mat
                    self._mark_material_used(material, best_attr['pos'], material_texts, used_materials)

                if material == 'Q235' and matched_idx_text:
                    # [v5.37z 修复2] 序号文字带#号，不应从中提取材质
                    idx_content = matched_idx_text.get('content', '')
                    if '#' not in idx_content and '＃' not in idx_content:
                        idx_mat = self.text_extractor.parse_material(idx_content)
                        if idx_mat:
                            material = idx_mat
                            self._mark_material_used(material, matched_idx_text['pos'], material_texts, used_materials)

                if material == 'Q235':
                    material = self._find_material(part_center, material_texts, part_diag, used_materials, attr_pos=best_attr.get('pos'))

                # 拼接标记
                sub_index = None
                is_original_sub = False  # [v5.37z8-修复] 标记来源

                if index_val != 'xx':
                    candidates = []
                    for idx, st in enumerate(split_texts):
                        if idx in used_splits:
                            continue
                        # 【v5.37z5 修复】跳过已被其他零件占用的 sub_index
                        if st.get('sub_index') in used_sub_indices.get((index_val, thickness), set()):
                            continue
                        if st.get('index') == index_val:
                            dist = math.hypot(st['pos'][0] - part_center[0], st['pos'][1] - part_center[1])
                            if dist < split_search_radius:
                                candidates.append((dist, idx, st['sub_index']))
                    if candidates:
                        candidates.sort()
                        _, best_idx, sub_index = candidates[0]
                        used_splits.add(best_idx)
                        is_original_sub = True  # 图纸原生
                    # 记录 split_texts 分配的 sub_index
                    if sub_index:
                        used_sub_indices[(index_val, thickness)].add(sub_index)

                if sub_index is None:
                    candidates = []
                    for idx, st in enumerate(split_texts):
                        if idx in used_splits:
                            continue
                        # 【v5.37z5 修复】跳过已被其他零件占用的 sub_index
                        if st.get('sub_index') in used_sub_indices.get((index_val, thickness), set()):
                            continue
                        if st.get('index') is None:
                            dist = math.hypot(st['pos'][0] - part_center[0], st['pos'][1] - part_center[1])
                            if dist < split_search_radius:
                                candidates.append((dist, idx, st['sub_index']))
                    if candidates:
                        candidates.sort()
                        _, best_idx, sub_index = candidates[0]
                        used_splits.add(best_idx)
                        is_original_sub = True  # 图纸原生
                    # 记录 split_texts 分配的 sub_index
                    if sub_index:
                        used_sub_indices[(index_val, thickness)].add(sub_index)

                # 【v5.37z5 修复】同序号同厚度自动递增 _01, _02, _03...
                if sub_index is None:
                    combo_key = (index_val, thickness)
                    candidate = 1
                    while f"{candidate:02d}" in used_sub_indices[combo_key]:
                        candidate += 1
                    sub_index = f"{candidate:02d}"
                    used_sub_indices[combo_key].add(sub_index)
                    is_original_sub = False  # 自动递增

                thickness_str = f"T{int(thickness)}"
                # [表面状态] 检测镀锌/喷漆/喷塑，文件夹名加后缀，文件名不变
                # [v5.37z8-文件夹结构] 材质/表面状态/厚度 三级结构，有则建无则跳过
                surface = self.text_extractor.parse_surface(best_attr.get('content', ''))
                if not surface and matched_idx_text:
                    surface = self.text_extractor.parse_surface(matched_idx_text.get('content', ''))

                material_str = material  # 如 Q235, 304
                dir_key = (material_str, thickness_str, surface)
                target_dir = thickness_dirs.get(dir_key)
                if not target_dir:
                    material_dir = output_dir / material_str
                    material_dir.mkdir(parents=True, exist_ok=True)
                    if surface:
                        surface_dir = material_dir / surface
                        surface_dir.mkdir(parents=True, exist_ok=True)
                        target_dir = surface_dir / thickness_str
                    else:
                        target_dir = material_dir / thickness_str
                    target_dir.mkdir(parents=True, exist_ok=True)
                    thickness_dirs[dir_key] = target_dir
                qty_str = str(qty) if has_qty else "1"

                if sub_index:
                    index_display = f"{index_val}#_{sub_index}"
                else:
                    index_display = f"{index_val}#"

                new_name = f"{self.full_prefix}_{index_display}_{material}_{thickness_str}_B_{process_code}_{qty_str}.dxf"
                print(f"  ✓ {thickness_str}/{new_name} (材质{material})")
                success += 1
                self.stats['matched'] += 1

            else:
                # [v5.37z3 保留] 无属性分支：不借用任何序号和厚度，直接进"其他"
                index_val = 'xx'
                thickness = 0
                process_code = 'xxxxx'
                material = 'Q235'

                target_dir = other_dir
                thickness_str = "T未知"

                new_name = f"{self.full_prefix}_P{i:02d}_{material}_{thickness_str}_B_{process_code}_1.dxf"
                print(f"  ⚠ 其他/{new_name} (无属性，不借用)")
                self.stats['unmatched'] += 1

            new_path = target_dir / new_name
            # [v5.37z8-修复] 区分原生/自动 sub_index 的冲突处理
            if new_path.exists() and index_val != 'xx' and thickness > 0:
                if is_original_sub:
                    # 图纸原生 sub_index 重复，扔"其他"
                    target_dir = other_dir
                    new_path = target_dir / new_name
                    print(f"  ⚠ 原生拼接标记重复，放其他: {new_name}")
                else:
                    # 自动递增 sub_index 重复，继续递增
                    base_sub = int(sub_index) if sub_index else 1
                    candidate = base_sub + 1
                    while candidate <= 99:
                        index_display = f"{index_val}#_{candidate:02d}"
                        new_name = f"{self.full_prefix}_{index_display}_{material}_{thickness_str}_B_{process_code}_{qty_str}.dxf"
                        new_path = target_dir / new_name
                        if not new_path.exists():
                            break
                        candidate += 1
                    if candidate <= 99:
                        print(f"  ⚠ 自动递增: {new_name}")
                    else:
                        print(f"  ⚠ 递增到99仍冲突，放其他: {new_name}")
            elif new_path.exists():
                print(f"  ⚠ 覆盖已存在文件: {new_name}")

            try:
                shutil.move(str(temp_path), str(new_path))
            except Exception as e:
                print(f"    ⚠ 移动失败: {e}")

        return success

    def _export_and_match(self, source_doc, parts, attr_texts, index_texts, process_texts, material_texts, split_texts, qty_texts, output_dir):
        if self.text_mode == 'loose':
            return self._export_and_match_loose(source_doc, parts, attr_texts, index_texts, process_texts, material_texts, split_texts, qty_texts, output_dir)
        return self._export_and_match_compact(source_doc, parts, attr_texts, index_texts, process_texts, material_texts, split_texts, qty_texts, output_dir)


    def _extract_drawing_no(self, input_path: Path):
        filename = input_path.stem
        clean = re.sub(r'[\\/*?:"<>|]', '', filename)

        fullwidth_map = {
            'Ａ': 'A', 'Ｂ': 'B', 'Ｃ': 'C', 'Ｄ': 'D', 'Ｅ': 'E',
            'Ｆ': 'F', 'Ｇ': 'G', 'Ｈ': 'H', 'Ｉ': 'I', 'Ｊ': 'J',
            'Ｋ': 'K', 'Ｌ': 'L', 'Ｍ': 'M', 'Ｎ': 'N', 'Ｏ': 'O',
            'Ｐ': 'P', 'Ｑ': 'Q', 'Ｒ': 'R', 'Ｓ': 'S', 'Ｔ': 'T',
            'Ｕ': 'U', 'Ｖ': 'V', 'Ｗ': 'W', 'Ｘ': 'X', 'Ｙ': 'Y', 'Ｚ': 'Z',
            'ａ': 'a', 'ｂ': 'b', 'ｃ': 'c', 'ｄ': 'd', 'ｅ': 'e',
            'ｆ': 'f', 'ｇ': 'g', 'ｈ': 'h', 'ｉ': 'i', 'ｊ': 'j',
            'ｋ': 'k', 'ｌ': 'l', 'ｍ': 'm', 'ｎ': 'n', 'ｏ': 'o',
            'ｐ': 'p', 'ｑ': 'q', 'ｒ': 'r', 'ｓ': 's', 'ｔ': 't',
            'ｕ': 'u', 'ｖ': 'v', 'ｗ': 'w', 'ｘ': 'x', 'ｙ': 'y', 'ｚ': 'z',
            '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
            '５': '5', '６': '6', '７': '7', '８': '8', '９': '9',
            '＿': '_', '－': '-', '　': ' ', '：': ':',
        }
        clean = ''.join(fullwidth_map.get(c, c) for c in clean)

        match = re.search(r'([A-Za-z]{1,4})[_\-\s]+[A-Za-z]*?(\d{6})(\d+)(?!\d)', clean)
        if match:
            priority = match.group(1).upper()
            drawing_no = match.group(2)
            piece_count = int(match.group(3))
            full_prefix = f"{priority}_{match.group(2)}{match.group(3)}"
            return drawing_no, piece_count, full_prefix

        match2 = re.search(r'[A-Za-z]*?(\d{6})(\d+)(?!\d)', clean)
        if match2:
            drawing_no = match2.group(1)
            piece_count = int(match2.group(2))
            full_prefix = f"A_{match2.group(1)}{match2.group(2)}"
            return drawing_no, piece_count, full_prefix

        match3 = re.search(r'(\d{6})(\d*)(?!\d)', clean)
        if match3:
            digits = match3.group(1)
            piece_str = match3.group(2)
            piece_count = int(piece_str) if piece_str else 1
            drawing_no = digits
            full_prefix = f"A_{digits}{piece_str if piece_str else '1'}"
            return drawing_no, piece_count, full_prefix

        match4 = re.search(r'(\d{4,5})(?!\d)', clean)
        if match4:
            digits = match4.group(1)
            if len(digits) == 4:
                drawing_no = digits + "01"
                piece_count = 1
            else:
                drawing_no = digits[:4] + "01"
                piece_count = int(digits[4:])
            full_prefix = f"A_{drawing_no}{piece_count}"
            return drawing_no, piece_count, full_prefix

        try:
            mtime = input_path.stat().st_mtime
            dt = datetime.fromtimestamp(mtime)
            drawing_no = dt.strftime("%m%d") + "01"
        except Exception:
            dt = datetime.now()
            drawing_no = dt.strftime("%m%d") + "01"

        piece_count = 1
        full_prefix = f"A_{drawing_no}1"
        return drawing_no, piece_count, full_prefix

    def _extract_all_entities(self, msp, doc):
        entities = []
        eid = 0

        if self.explode_blocks:
            for insert in msp.query('INSERT'):
                # 【v5.37x-INSERT-fix】保留 INSERT 本身，防止 explode 后游离实体丢失
                insert_info = self._get_entity_info(insert, eid, 'INSERT')
                if insert_info:
                    entities.append(insert_info)
                    eid += 1
                
                # 【v5.37z8 修复】SolidWorks 中心标记符号只保留块参照本身，不 explode，
                # 防止十字标被拆分为大量独立线段形成孤立的零件组
                block_name = getattr(insert.dxf, 'name', '').upper()
                if 'CENTERMARK' in block_name:
                    continue
                
                for ent in self._explode_insert(insert, doc):
                    info = self._get_entity_info(ent, eid)
                    if info and info['type'] not in ['TEXT', 'MTEXT']:
                        entities.append(info)
                        eid += 1

        for etype in ['LWPOLYLINE', 'LINE', 'ARC', 'CIRCLE', 'ELLIPSE', 'SPLINE']:
            for e in msp.query(etype):
                info = self._get_entity_info(e, eid, etype)
                if info:
                    entities.append(info)
                    eid += 1

        return entities

    def _explode_insert(self, insert, doc):
        exploded = []
        try:
            block = doc.blocks.get(insert.dxf.name)
            if not block:
                return exploded

            transform = Matrix44.chain(
                Matrix44.translate(insert.dxf.insert[0], insert.dxf.insert[1], 0),
                Matrix44.z_rotate(math.radians(insert.dxf.rotation)),
                Matrix44.scale(insert.dxf.xscale, insert.dxf.yscale, 1)
            )

            for entity in block:
                try:
                    copied = entity.copy()
                    copied.transform(transform)
                    exploded.append(copied)
                except Exception:
                    continue
        except Exception:
            pass

        return exploded

    def _get_entity_info(self, entity, eid, etype=None):
        if etype is None:
            etype = entity.dxftype()

        try:
            if etype == 'LWPOLYLINE':
                points = list(entity.vertices_in_wcs())
                xs = [p[0] for p in points]
                ys = [p[1] for p in points]
                bounds = (min(xs), min(ys), max(xs), max(ys))
                return {
                    'id': eid, 'type': etype, 'entity': entity,
                    'bounds': bounds,
                    'center': ((bounds[0]+bounds[2])/2, (bounds[1]+bounds[3])/2),
                    'endpoints': self._get_endpoints(points, entity.closed, bounds),
                }

            elif etype == 'LINE':
                s = entity.dxf.start
                e = entity.dxf.end
                bounds = (min(s[0], e[0]), min(s[1], e[1]), max(s[0], e[0]), max(s[1], e[1]))
                return {
                    'id': eid, 'type': etype, 'entity': entity,
                    'bounds': bounds,
                    'center': ((s[0]+e[0])/2, (s[1]+e[1])/2),
                    'endpoints': [(s[0], s[1]), (e[0], e[1])],
                }

            elif etype == 'ARC':
                c = entity.dxf.center
                r = entity.dxf.radius
                sa = math.radians(entity.dxf.start_angle)
                ea = math.radians(entity.dxf.end_angle)
                x1, y1 = c[0] + r*math.cos(sa), c[1] + r*math.sin(sa)
                x2, y2 = c[0] + r*math.cos(ea), c[1] + r*math.sin(ea)
                angles = [sa, ea] + [a for a in [0, math.pi/2, math.pi, 3*math.pi/2]
                                     if self._angle_between(sa, ea, a)]
                xs = [c[0] + r*math.cos(a) for a in angles]
                ys = [c[1] + r*math.sin(a) for a in angles]
                bounds = (min(xs), min(ys), max(xs), max(ys))
                return {
                    'id': eid, 'type': etype, 'entity': entity,
                    'bounds': bounds, 'center': (c[0], c[1]),
                    'endpoints': [(x1, y1), (x2, y2)],
                }

            elif etype == 'CIRCLE':
                c = entity.dxf.center
                r = entity.dxf.radius
                bounds = (c[0]-r, c[1]-r, c[0]+r, c[1]+r)
                return {
                    'id': eid, 'type': etype, 'entity': entity,
                    'bounds': bounds, 'center': (c[0], c[1]),
                    'endpoints': [(c[0]-r, c[1]), (c[0]+r, c[1]), (c[0], c[1]-r), (c[0], c[1]+r), (c[0], c[1])],
                }

            # 【v5.37x-INSERT-fix】增加 INSERT 块参照提取（SolidWorks中心标记等）
            elif etype == 'INSERT':
                pos = entity.dxf.insert
                return {
                    'id': eid, 'type': etype, 'entity': entity,
                    'bounds': (pos[0], pos[1], pos[0], pos[1]),
                    'center': (pos[0], pos[1]),
                    'endpoints': [(pos[0], pos[1])],
                }

        except Exception:
            return None

    def _get_endpoints(self, points, is_closed, bounds=None):
        if len(points) < 2:
            return []
        if is_closed and bounds:
            min_x, min_y, max_x, max_y = bounds
            return [(min_x, min_y), (max_x, min_y), (min_x, max_y), (max_x, max_y),
                    ((min_x+max_x)/2, (min_y+max_y)/2)]
        return [(points[0][0], points[0][1]), (points[-1][0], points[-1][1])]

    def _angle_between(self, start, end, angle):
        if start <= end:
            return start <= angle <= end
        return angle >= start or angle <= end

    def _build_connection_groups(self, entities):
        if not entities:
            return []

        n = len(entities)
        parent = list(range(n))

        def find(x):
            if parent[x] != x:
                parent[x] = find(parent[x])
            return parent[x]

        def union(x, y):
            px, py = find(x), find(y)
            if px != py:
                parent[px] = py

        point_map = defaultdict(list)
        tol = self.gap_tolerance

        for i, ent in enumerate(entities):
            for p in ent.get('endpoints', []):
                key = (round(p[0]/tol), round(p[1]/tol))
                point_map[key].append((i, p))

        for i, ent in enumerate(entities):
            for p in ent.get('endpoints', []):
                key = (round(p[0]/tol), round(p[1]/tol))
                for dx in [-1, 0, 1]:
                    for dy in [-1, 0, 1]:
                        for j, other_p in point_map.get((key[0]+dx, key[1]+dy), []):
                            if i != j and math.hypot(p[0]-other_p[0], p[1]-other_p[1]) < tol:
                                union(i, j)

        groups_map = defaultdict(list)
        for i, ent in enumerate(entities):
            groups_map[find(i)].append(ent)

        return list(groups_map.values())

    def _cluster_by_proximity(self, groups):
        if not groups:
            return []

        group_data = []
        for group in groups:
            if not group:
                continue
            all_bounds = [e['bounds'] for e in group]
            min_x = min(b[0] for b in all_bounds)
            min_y = min(b[1] for b in all_bounds)
            max_x = max(b[2] for b in all_bounds)
            max_y = max(b[3] for b in all_bounds)
            group_data.append({
                'entities': group,
                'bounds': (min_x, min_y, max_x, max_y),
                'center': ((min_x+max_x)/2, (min_y+max_y)/2),
                'merged': False,
            })

        margin = self.cluster_margin
        n = len(group_data)

        for i in range(n):
            if group_data[i]['merged']:
                continue
            merged_any = True
            while merged_any:
                merged_any = False
                cb = group_data[i]['bounds']
                for j in range(i+1, n):
                    if group_data[j]['merged']:
                        continue
                    ob = group_data[j]['bounds']
                    if not (cb[2] < ob[0]-margin or ob[2] < cb[0]-margin or
                            cb[3] < ob[1]-margin or ob[3] < cb[1]-margin):
                        group_data[i]['entities'].extend(group_data[j]['entities'])
                        group_data[i]['bounds'] = (min(cb[0], ob[0]), min(cb[1], ob[1]),
                                                   max(cb[2], ob[2]), max(cb[3], ob[3]))
                        group_data[i]['center'] = ((group_data[i]['bounds'][0]+group_data[i]['bounds'][2])/2,
                                                   (group_data[i]['bounds'][1]+group_data[i]['bounds'][3])/2)
                        group_data[j]['merged'] = True
                        merged_any = True

        return [{'entities': g['entities'], 'bounds': g['bounds'], 'center': g['center']}
                for g in group_data if not g['merged']]

    def _get_diagonal_size(self, bounds):
        return math.hypot(bounds[2]-bounds[0], bounds[3]-bounds[1])

    def _export_part_temp(self, source_doc, part, temp_name, output_dir):
        try:
            target_doc = ezdxf.new('R2010')
            target_msp = target_doc.modelspace()
            importer = Importer(source_doc, target_doc)
            entities_to_import = [e['entity'] for e in part['entities']]
            if not entities_to_import:
                return None
            importer.import_entities(entities_to_import, target_msp)
            importer.finalize()

            original_center = part['center']
            dx, dy = -original_center[0], -original_center[1]
            for entity in list(target_msp):
                try:
                    entity.transform(Matrix44.translate(dx, dy, 0))
                except Exception:
                    pass

            temp_path = output_dir / temp_name
            target_doc.saveas(str(temp_path))
            return {'path': temp_path, 'center': original_center, 'bounds': part['bounds']}

        except Exception as e:
            logger.error(f"导出临时文件失败: {e}")
            return None

    def _generate_batch_import_excel(self, output_dir: Path):
        cypnest_batchimport = output_dir

        total_piece_count = getattr(self, 'piece_count', 1)
        thickness_records = {}

        # [v5.37z8-文件夹结构] 递归遍历所有 DXF，按 材质_厚度 分组（不区分表面状态）
        for dxf_file in output_dir.rglob("*.dxf"):
            if dxf_file.parent.name == "外框":
                continue

            rel_path = dxf_file.relative_to(output_dir)
            parts = rel_path.parts

            # 新结构: 材质/表面状态/厚度/文件名.dxf (4层) 或 材质/厚度/文件名.dxf (3层)
            # 或 其他/文件名.dxf (2层)
            material = ""
            thickness = "T未知"
            if len(parts) >= 3:
                material = parts[0]  # 第一层 = 材质
                # 判断第二层是表面状态还是厚度
                if len(parts) >= 4:
                    # 4层: 材质/表面状态/厚度/文件
                    thickness = parts[2]
                else:
                    # 3层: 材质/厚度/文件
                    thickness = parts[1]
            elif len(parts) == 2:
                if parts[0] == "其他":
                    material = "其他"
                    thickness = "T未知"
                else:
                    material = parts[0]
                    thickness = "T未知"

            folder_key = f"{material}_{thickness}" if material else thickness
            if folder_key not in thickness_records:
                thickness_records[folder_key] = []

            stem = dxf_file.stem

            single_piece_qty = 1
            match = re.search(r'_(\d+)x(\d+)$', stem)
            if match:
                single_piece_qty = int(match.group(1))
            else:
                match2 = re.search(r'_(\d+)$', stem)
                if match2:
                    single_piece_qty = int(match2.group(1))

            amount = total_piece_count * single_piece_qty

            thickness_records[folder_key].append({
                'PartName': stem,
                'Amount': amount,
                'DXFFilePath': str(dxf_file.absolute()),
            })

        if not thickness_records:
            print(f"  ⚠ 未找到可导入的 DXF 文件")
            return

        for thickness, records in thickness_records.items():
            # [v5.37z6-修复15] 其他文件夹不生成Excel
            if '其他' in str(thickness):
                continue
            if not records:
                continue

            records.sort(key=lambda x: x['PartName'])

            excel_path = cypnest_batchimport / f"{self.drawing_no}_{thickness}_BatchImport.xlsx"
            csv_path = cypnest_batchimport / f"{self.drawing_no}_{thickness}_BatchImport.csv"

            excel_success = False
            if OPENPYXL_AVAILABLE:
                try:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "PartsDefinition"

                    headers = ['PartName', 'Amount', 'DXFFilePath']
                    ws.append(headers)

                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF", size=11)
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border

                    for record in records:
                        ws.append([record['PartName'], record['Amount'], record['DXFFilePath']])

                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        for cell in row:
                            cell.border = thin_border
                            cell.alignment = Alignment(vertical='center')
                            if cell.column == 2:
                                cell.alignment = Alignment(horizontal='center', vertical='center')

                    ws.column_dimensions['A'].width = 35
                    ws.column_dimensions['B'].width = 12
                    ws.column_dimensions['C'].width = 80

                    wb.save(str(excel_path))
                    print(f"  📋 {thickness} BatchImport Excel 已生成: {excel_path.name}")
                    print(f"     图号件数: {total_piece_count}, 共 {len(records)} 条记录")
                    excel_success = True
                except Exception as e:
                    logger.warning(f"{thickness} Excel 生成失败: {e}，回退到 CSV")

            if not excel_success:
                import csv
                with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(['PartName', 'Amount', 'DXFFilePath'])
                    for record in records:
                        writer.writerow([record['PartName'], record['Amount'], record['DXFFilePath']])
                print(f"  📋 {thickness} BatchImport CSV 已生成: {csv_path.name}")
                print(f"     图号件数: {total_piece_count}, 共 {len(records)} 条记录")


class BatchProcessor:
    def __init__(self, splitter):
        self.splitter = splitter

    def _extract_archive(self, archive_path: Path) -> Optional[Path]:
        suffix = archive_path.suffix.lower()
        temp_dir = Path(tempfile.mkdtemp(prefix="dxf_split_"))

        try:
            if suffix == '.zip':
                with zipfile.ZipFile(archive_path, 'r') as zf:
                    zf.extractall(temp_dir)
                return temp_dir
            elif suffix == '.rar' and RAR_AVAILABLE:
                with rarfile.RarFile(archive_path, 'r') as rf:
                    rf.extractall(temp_dir)
                return temp_dir
            else:
                logger.warning(f"不支持的压缩格式或缺少依赖: {suffix}")
                shutil.rmtree(temp_dir, ignore_errors=True)
                return None
        except Exception as e:
            logger.error(f"解压失败 {archive_path}: {e}")
            shutil.rmtree(temp_dir, ignore_errors=True)
            return None

    def find_dxf_files(self, directory, recursive=True):
        pattern = "**/*.dxf" if recursive else "*.dxf"
        files = [f for f in Path(directory).glob(pattern) if "拆分结果" not in str(f)]
        files.sort()
        return files

    def process_directory(self, input_dir, output_base_dir, recursive=True):
        stats = {'total_files': 0, 'processed_files': 0, 'total_parts': 0, 'frames': 0, 'errors': []}
        dxf_files = self.find_dxf_files(Path(input_dir), recursive)

        archive_patterns = ["**/*.zip", "**/*.rar"] if recursive else ["*.zip", "*.rar"]
        archives = []
        for ap in archive_patterns:
            archives.extend([f for f in Path(input_dir).glob(ap) if "拆分结果" not in str(f)])
        archives.sort()

        if not dxf_files and not archives:
            print(f"\n⚠ 未找到DXF文件或压缩包")
            return stats

        print(f"\n📁 找到 {len(dxf_files)} 个DXF文件, {len(archives)} 个压缩包")

        for dxf_file in dxf_files:
            base_name = f"{dxf_file.stem}_拆分结果"
            output_dir = Path(output_base_dir) / base_name
            # [v5.37z8-修复] 直接覆盖旧拆分结果，避免桌面多个编号文件夹混淆
            if output_dir.exists():
                shutil.rmtree(output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
            parts_count = self.splitter.process_file(dxf_file, output_dir)
            stats['total_files'] += 1
            if parts_count > 0:
                stats['processed_files'] += 1
                stats['total_parts'] += parts_count
            stats['frames'] += self.splitter.stats.get('frames', 0)

        for archive in archives:
            print(f"\n📦 处理压缩包: {archive.name}")
            temp_dir = self._extract_archive(archive)
            if not temp_dir:
                stats['errors'].append(f"无法解压: {archive.name}")
                continue

            try:
                sub_stats = self.process_directory(temp_dir, output_base_dir, recursive=True)
                stats['total_files'] += 1
                if sub_stats['processed_files'] > 0:
                    stats['processed_files'] += 1
                stats['total_parts'] += sub_stats['total_parts']
                stats['frames'] += sub_stats.get('frames', 0)
                if sub_stats['errors']:
                    stats['errors'].extend([f"[{archive.name}] {e}" for e in sub_stats['errors']])
            finally:
                shutil.rmtree(temp_dir, ignore_errors=True)

        return stats

    def process_path(self, input_path, output_base_dir):
        input_path = Path(input_path)
        if not input_path.exists():
            return {'total_files': 0, 'processed_files': 0, 'total_parts': 0, 'frames': 0, 'errors': ['路径不存在']}

        if input_path.is_file() and input_path.suffix.lower() in ['.zip', '.rar']:
            print(f"\n📦 直接处理压缩包: {input_path.name}")
            temp_dir = self._extract_archive(input_path)
            if not temp_dir:
                return {'total_files': 1, 'processed_files': 0, 'total_parts': 0, 'frames': 0, 
                        'errors': [f'解压失败: {input_path.name}']}
            try:
                result = self.process_directory(temp_dir, output_base_dir, recursive=True)
                result['total_files'] = 1
                return result
            finally:
                shutil.rmtree(temp_dir, ignore_errors=True)

        if input_path.is_file() and input_path.suffix.lower() == '.dxf':
            base_name = f"{input_path.stem}_拆分结果"
            output_dir = Path(output_base_dir) / base_name
            # [v5.37z8-修复] 直接覆盖旧拆分结果，避免桌面多个编号文件夹混淆
            if output_dir.exists():
                shutil.rmtree(output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
            parts_count = self.splitter.process_file(input_path, output_dir)
            return {
                'total_files': 1, 
                'processed_files': 1 if parts_count > 0 else 0,
                'total_parts': parts_count, 
                'frames': self.splitter.stats.get('frames', 0), 
                'errors': []
            }

        return self.process_directory(input_path, output_base_dir, recursive=True)


def print_summary(stats):
    print(f'\n{"="*60}')
    print("📊 处理汇总")
    print(f"  总文件数: {stats['total_files']}")
    print(f"  成功处理: {stats['processed_files']}")
    print(f"  导出零件数: {stats['total_parts']}")
    print(f"  剥离外框数: {stats.get('frames', 0)}")
    if stats.get('errors'):
        print(f"  错误/警告: {len(stats['errors'])} 条")
        for e in stats['errors'][:5]:
            print(f"    ⚠ {e}")
    print('='*60)


def main():
    print('='*60)
    print("DXF批量拆分器 v5.37z8 - 双模式文字提取版")
    print("模式1[紧凑精确]: 规范制图，支持拼接标记如7#-2T4*88")
    print("模式2[宽松覆盖]: 钢模面板/大长条，大搜索半径5000mm")
    print('='*60)

    try:
        import tkinter as tk
        from tkinter import filedialog
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()

        file_paths = filedialog.askopenfilenames(
            title="选择DXF/ZIP/RAR文件（Ctrl+A全选，Ctrl点选多个）",
            filetypes=[
                ("所有文件", "*.*"),
                ("DXF文件", "*.dxf"),
                ("ZIP压缩包", "*.zip"),
                ("RAR压缩包", "*.rar"),
            ]
        )

        if not file_paths:
            print("未选择文件，退出")
            return

        last_output = load_last_output()
        initial_dir = last_output if last_output and Path(last_output).exists() else str(Path.home() / "Desktop")

        output_dir = filedialog.askdirectory(title="选择输出文件夹", initialdir=initial_dir)
        if not output_dir:
            output_dir = str(Path.home() / "Desktop" / "DXF拆分结果")
        else:
            save_last_output(output_dir)

        use_gui = True

    except Exception as e:
        logger.warning(f"GUI初始化失败: {e}")
        print(f"⚠ GUI不可用，使用命令行模式")
        raw = input("文件路径（多个用逗号分隔）: ").strip().strip('"')
        file_paths = [p.strip().strip('"') for p in raw.split(",") if p.strip()]
        output_dir = input("输出文件夹路径(默认桌面): ").strip().strip('"') or str(Path.home() / "Desktop" / "DXF拆分结果")
        use_gui = False

    if not file_paths:
        print("未指定输入，退出")
        return

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    cfg = load_splitter_config()
    splitter = UniversalPartSplitter(
        gap_tolerance=cfg.get('gap_tolerance', 0.5),
        cluster_margin=cfg.get('cluster_margin', 5.0),
        min_part_size=cfg.get('min_part_size', 5.0),
        config=cfg
    )
    splitter.select_parameters_interactive()
    processor = BatchProcessor(splitter)

    total_stats = {'total_files': 0, 'processed_files': 0, 'total_parts': 0, 'frames': 0, 'errors': []}

    for file_path in file_paths:
        print(f"\n{'='*60}")
        print(f"📂 处理: {Path(file_path).name}")
        print(f"{'='*60}")

        stats = processor.process_path(Path(file_path), Path(output_dir))

        total_stats['total_files'] += stats['total_files']
        total_stats['processed_files'] += stats['processed_files']
        total_stats['total_parts'] += stats['total_parts']
        total_stats['frames'] += stats.get('frames', 0)
        if stats.get('errors'):
            total_stats['errors'].extend(stats['errors'])

    print_summary(total_stats)

    if use_gui:
        try:
            messagebox.showinfo("处理完成", 
                f"成功: {total_stats['processed_files']}/{total_stats['total_files']}\n"
                f"零件: {total_stats['total_parts']}个\n"
                f"外框: {total_stats['frames']}个")
        except Exception:
            pass


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.critical(f"程序崩溃: {e}")
        logger.critical(traceback.format_exc())
        print(f"\n{'='*60}")
        print("程序发生严重错误")
        print(f"错误: {e}")
        print('='*60)
        input("按回车键退出...")